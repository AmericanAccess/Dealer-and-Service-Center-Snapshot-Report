"""AA Weekly Leadership Report — built to the AA Style & Formatting Specification.

Eval = RED, Install = BLUE.  East = blue, Central = red, West = silver.
FOIA tab dropped per Micah, 6 Aug 2026.

IMPORTANT: do NOT run a LibreOffice recalc pass on the output — it strips the
conditional-formatting data bars. There are no live formulas, so it opens clean.

    AA_EXPORT=... AA_WEEK=2026-07-27 AA_LOGO=aa_logo.png python3 build_weekly_report.py
"""
import os, re, warnings
warnings.filterwarnings('ignore')
import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import LineChart, BarChart, AreaChart, PieChart, DoughnutChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.marker import DataPoint as _DP2
from openpyxl.drawing.colors import SchemeColor, ColorChoice
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (RichTextProperties, Paragraph, ParagraphProperties,
                                   CharacterProperties)
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.data_source import StrRef, AxDataSource
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter as L

EXPORT = os.environ.get('AA_EXPORT', '/mnt/user-data/uploads/AA_Weekly_Snapshot_Report_8-6-2026_9-04-30_AM.xlsx')
WEEK   = pd.Timestamp(os.environ.get('AA_WEEK', '2026-07-27'))
LOGO   = os.environ.get('AA_LOGO', '/mnt/user-data/outputs/aa_logo.png')
OUT    = os.environ.get('AA_OUT', f'/mnt/user-data/outputs/AA_Weekly_Report_{WEEK.date()}.xlsx')
WOW_N  = 5

# ---------------- §1 brand palette ----------------
AA_BLUE='004890'; AA_RED='D2001A'; AA_SILVER='A6A6A6'; AA_BLUE_DK='003462'
LRED1='F4B7BE'; LRED2='FBDDE1'; LRED3='FCE4E6'
LBLU1='BDD3EA'; LBLU2='D6E4F2'; LBLU3='DDEBF7'
LBLGRY='F2F2F2'; CHAN_GREEN='D9EAD3'; CHAN_AMBER='FCE4E6'
KPI_G='C6E0B4'; KPI_A='FFE699'; KPI_R='F4B7BE'
SUBHDR='6B7280'; NOUP='C00000'
BRD_T='D9D9D9'; BRD_I='BFBFBF'

EAST={1,2,3,4,5,6,7,8}; CENTRAL={9,10,11,12,15,16,23}; WEST={17,19,20,21,22}
VAMC_OVERRIDE={'IOWA CITY VA':23,'VAMC SEPULVEDA CA':22}
REG_COLOR={'East':AA_BLUE,'Central':AA_RED,'West':AA_SILVER}

# ---------------- §2 fonts & styles ----------------
FN='Calibri'
def F(sz=11,b=False,c='000000',i=False): return Font(name=FN,size=sz,bold=b,color=c,italic=i)
TITLE16=F(16,True,AA_BLUE); TITLE15=F(15,True,AA_BLUE)
SUBT=F(10,False,SUBHDR,i=True)
HDW=F(11,True,'FFFFFF'); BD=F(11,True); RG=F(11); SM=F(10)
NOTE=F(9,False,SUBHDR,i=True)
def P(hex_): return PatternFill('solid',fgColor=hex_)
FBLUE=P(AA_BLUE); FSUB=P(SUBHDR); FLBL=P(LBLGRY)
thin=Side(style='thin',color=BRD_T)
BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal='center',vertical='center')
LFT=Alignment(horizontal='left',vertical='center')
WRAP=Alignment(horizontal='left',vertical='top',wrap_text=True)

# ---------------- data ----------------
df=pd.read_excel(EXPORT,sheet_name=0)
df['code']=df['Cancellation Code'].astype(str).str.split('-').str[0].str.strip()
df.loc[df['Cancellation Code'].isna(),'code']=None
a=df[df['code'].isna()].copy()

# Safety net: Dataverse has been seen to reuse a Project ID across two entirely
# different projects (different veteran, different VAMC, different technician).
# When that happens and both rows survive the cancellation filter, every count
# doubles for one of them. Keep the first surviving row per ID and report the
# rest rather than silently either double-counting or dropping data unnoticed.
_dupmask = a['Project ID'].duplicated(keep=False)
DUP_PROJECT_IDS = sorted(a.loc[_dupmask, 'Project ID'].unique().tolist())
if DUP_PROJECT_IDS:
    print(f'WARNING: {len(DUP_PROJECT_IDS)} Project ID(s) reused across distinct rows '
          f'after the cancellation filter, keeping first: {DUP_PROJECT_IDS}')
a = a.drop_duplicates(subset='Project ID', keep='first')

def visn_of(c):
    if not isinstance(c,str): return None
    c=c.strip()
    m=re.search(r'#\s*0*(\d{1,2})',c)
    if m: return int(m.group(1))
    m=re.search(r'VISN\s*0*(\d{1,2})',c,re.I)
    if m: return int(m.group(1))
    return VAMC_OVERRIDE.get(c.upper())
def region_of(v):
    return 'East' if v in EAST else 'Central' if v in CENTRAL else 'West' if v in WEST else 'Unassigned'
def wk_of(t): return (t-pd.Timedelta(days=t.weekday())).normalize() if pd.notna(t) else pd.NaT

# ---------------------------------------------------------------------------
# AA HOLIDAY / DAY-COUNT ENGINE
# Implements the GAP Late-Codes framework's non-working-day rules verbatim.
# Governing text (Late-Codes doc, SECTION PARAMETERS & CLOCK ANCHORS):
#   "NON-WORKING DAYS -- the 6 AA holidays, skipped in EVERY mode."
#   "Observed-day rule applies to the three FIXED-DATE holidays only (New Year,
#    Jul 4, Christmas): falls on a Saturday -> AA closed the preceding Friday;
#    falls on a Sunday -> the following Monday. The three floating holidays are
#    always weekdays and never shift."
#   "Calendar mode skips holidays ONLY. It does NOT skip weekends ... So in the
#    three calendar sections a red-zone day can land on a Saturday or Sunday."
#   "Holidays are skipped EVERYWHERE -- the displayed day-count, the on-time
#    window, and the red-zone / flag positions all use the same holiday-adjusted
#    count, so every number in a string agrees."
# NOTE: the framework says an authoritative engine exists (daycount.py,
# sections.py) and that "the ENGINE WINS" on any disagreement. Those files are
# not available here, so this is implemented from the doc's own stated values
# (its Table 0 / Table 1). If the real engine ever becomes available, call it
# instead of this.
def aa_holidays(year):
    """The 6 AA holidays for a year, with the observed-day shift applied to the
    three fixed-date ones only."""
    out=[]
    for m,d in ((1,1),(7,4),(12,25)):                 # fixed-date, observed
        h=pd.Timestamp(year=year,month=m,day=d)
        if h.weekday()==5: h-=pd.Timedelta(days=1)     # Sat -> preceding Fri
        elif h.weekday()==6: h+=pd.Timedelta(days=1)   # Sun -> following Mon
        out.append(h.normalize())
    # floating: always weekdays, never shift
    may=pd.date_range(f'{year}-05-01',f'{year}-05-31',freq='D')
    out.append(max(d for d in may if d.weekday()==0).normalize())          # last Mon in May
    sep=pd.date_range(f'{year}-09-01',f'{year}-09-30',freq='D')
    out.append(min(d for d in sep if d.weekday()==0).normalize())          # first Mon in Sep
    nov=[d for d in pd.date_range(f'{year}-11-01',f'{year}-11-30',freq='D') if d.weekday()==3]
    out.append(nov[3].normalize())                                          # 4th Thu in Nov
    return set(out)

_HOL_CACHE={}
def _hols_for(a,b):
    yrs=range(min(a.year,b.year), max(a.year,b.year)+1)
    key=tuple(yrs)
    if key not in _HOL_CACHE:
        s=set()
        for y in yrs: s|=aa_holidays(y)
        _HOL_CACHE[key]=s
    return _HOL_CACHE[key]

def day_count(anchor,target,mode='calendar'):
    """Days from anchor to target. 'calendar' skips AA holidays only (NOT
    weekends). 'business' skips weekends AND AA holidays. Returns None if
    either date is missing. Vectorized via numpy busday_count -- weekmask
    '1111111' makes it count every weekday including weekends, so the only
    exclusion in calendar mode is the holiday list, exactly as the framework
    specifies."""
    if pd.isna(anchor) or pd.isna(target): return None
    a0=pd.Timestamp(anchor).normalize(); b0=pd.Timestamp(target).normalize()
    hols=sorted(_hols_for(min(a0,b0),max(a0,b0)))
    hol_np=np.array([h.date() for h in hols],dtype='datetime64[D]')
    wm='1111111' if mode=='calendar' else '1111100'
    return int(np.busday_count(a0.date(),b0.date(),weekmask=wm,holidays=hol_np))

def day_count_series(anchors,targets,mode='calendar'):
    """Vectorized day_count over two datetime Series. NaT-safe."""
    a=pd.to_datetime(anchors,errors='coerce'); b=pd.to_datetime(targets,errors='coerce')
    ok=a.notna()&b.notna()
    out=pd.Series([None]*len(a),index=a.index,dtype='object')
    if not ok.any(): return out
    lo=min(a[ok].min(),b[ok].min()); hi=max(a[ok].max(),b[ok].max())
    hols=sorted(_hols_for(lo,hi))
    hol_np=np.array([h.date() for h in hols],dtype='datetime64[D]')
    wm='1111111' if mode=='calendar' else '1111100'
    av=a[ok].dt.normalize().values.astype('datetime64[D]')
    bv=b[ok].dt.normalize().values.astype('datetime64[D]')
    out.loc[ok]=np.busday_count(av,bv,weekmask=wm,holidays=hol_np)
    return out

def clean_note(v, maxlen=200):
    """Timeline Note comes from a rich-text CKEditor field: raw HTML, embedded
    <img> tags, inline CSS. Strip it to plain, readable text."""
    import re, html
    s = str(v or '')
    if not s or s.lower() == 'nan': return ''
    s = re.sub(r'<img[^>]*>', ' [image] ', s, flags=re.I)   # note an image was there, drop the tag
    s = re.sub(r'<[^>]+>', ' ', s)                            # strip all remaining tags
    s = html.unescape(s)                                      # &nbsp; etc.
    s = re.sub(r'\s+', ' ', s).strip()
    return s[:maxlen]

def soften(hexcolor, amt=0.45):
    """Blend a hex color toward white so pie/donut slices read as pastel rather
    than saturated -- same hue, easier on the eye across many small slices."""
    r=int(hexcolor[0:2],16); g=int(hexcolor[2:4],16); b=int(hexcolor[4:6],16)
    r=int(r+(255-r)*amt); g=int(g+(255-g)*amt); b=int(b+(255-b)*amt)
    return f'{r:02X}{g:02X}{b:02X}'

def short(n):
    """AA Maryland Service Center -> AA Maryland SC;  Next Day Access Tampa -> NDA Tampa"""
    t=str(n)
    t=re.sub(r'\bNext Day Access\b','NDA',t,flags=re.I)
    t=re.sub(r'\bService Center\b','SC',t,flags=re.I)
    t=re.sub(r',?\s*LLC\b','',t,flags=re.I)
    t=re.sub(r'\s*-\s*',' ',t)
    return re.sub(r'\s+',' ',t).strip()

def is_sc(n): return bool(re.search(r'service\s+center',str(n),re.I))

a['region']=a['Customer'].apply(visn_of).apply(region_of)

# ---- optional: real document-upload dates from get_document_upload_dates ----
# When present, completion switches from the Quote-Sent/Packet-Sent proxy to the
# CREW'S OWN upload timestamp -- the fair completion date. Off by default so a
# plain run never silently changes what "completed" means.
UPLOAD_CSV = os.environ.get('AA_UPLOAD_CSV', '')
USE_REAL_UPLOAD = os.environ.get('AA_USE_UPLOAD_DATES', '0') == '1'
upl = None
if UPLOAD_CSV and os.path.exists(UPLOAD_CSV):
    upl = pd.read_csv(UPLOAD_CSV, dtype=str)
    for c in ('eval_upload','install_upload','eval_signoff','install_signoff'):
        upl[c] = pd.to_datetime(upl[c], errors='coerce', utc=True)
    for c in ('multiple_signoffs','cancellation_evidence'):
        upl[c] = upl[c].astype(str).str.strip().str.lower().eq('true')
    upl = upl.set_index('project')

TZMAP = {'CA':'America/Los_Angeles','NV':'America/Los_Angeles','OR':'America/Los_Angeles',
         'WA':'America/Los_Angeles','AZ':'America/Phoenix','CO':'America/Denver',
         'ID':'America/Denver','MT':'America/Denver','NM':'America/Denver','UT':'America/Denver',
         'WY':'America/Denver','AL':'America/Chicago','AR':'America/Chicago','IA':'America/Chicago',
         'IL':'America/Chicago','LA':'America/Chicago','MO':'America/Chicago','MS':'America/Chicago',
         'ND':'America/Chicago','OK':'America/Chicago','TN':'America/Chicago','TX':'America/Chicago',
         'WI':'America/Chicago','KS':'America/Chicago','MN':'America/Chicago','NE':'America/Chicago',
         'SD':'America/Chicago','CT':'America/New_York','DC':'America/New_York','FL':'America/New_York',
         'GA':'America/New_York','IN':'America/New_York','MA':'America/New_York','MD':'America/New_York',
         'ME':'America/New_York','MI':'America/New_York','NC':'America/New_York','NY':'America/New_York',
         'OH':'America/New_York','PA':'America/New_York','SC':'America/New_York','VA':'America/New_York',
         'VT':'America/New_York','WV':'America/New_York','NJ':'America/New_York','NH':'America/New_York',
         'RI':'America/New_York','DE':'America/New_York','KY':'America/New_York','PR':'America/Puerto_Rico',
         'AK':'America/Anchorage'}

def local_naive(ts, state):
    if pd.isna(ts): return pd.NaT
    tz = TZMAP.get(str(state), 'America/New_York')
    return ts.tz_convert(tz).tz_localize(None)

if upl is not None:
    a['_pid'] = a['Project ID'].astype(str)
    j = a['_pid'].map(upl['eval_upload']) if '_pid' in a else None
    a['eu_raw']  = a['_pid'].map(upl['eval_upload'])    if upl is not None else pd.NaT
    a['iu_raw']  = a['_pid'].map(upl['install_upload']) if upl is not None else pd.NaT
    a['es_raw']  = a['_pid'].map(upl['eval_signoff'])
    a['is_raw']  = a['_pid'].map(upl['install_signoff'])
    a['_ustate'] = a['_pid'].map(upl['state']) if 'state' in upl.columns else a['Customer'].apply(visn_of)
    a['msig']    = a['_pid'].map(upl['multiple_signoffs']).fillna(False)
    a['cevid']   = a['_pid'].map(upl['cancellation_evidence']).fillna(False)
    a['fstatus'] = a['_pid'].map(upl['folder_status']) if 'folder_status' in upl.columns else pd.Series(dtype=object)
    a['eup_local'] = [local_naive(t,s) for t,s in zip(a['eu_raw'], a['_ustate'])]
    a['iup_local'] = [local_naive(t,s) for t,s in zip(a['iu_raw'], a['_ustate'])]

a['edt']=pd.to_datetime(a['Eval Date and Time'],errors='coerce')
a['idt']=pd.to_datetime(a['Install Date and Time'],errors='coerce')
a['qsd']=pd.to_datetime(a['Quote Sent Date'],errors='coerce')
a['ewk']=a['edt'].apply(wk_of); a['iwk']=a['idt'].apply(wk_of)
# §6 completion definitions
# NOTE: the upload CSV was pulled by dateField=cw_installdateandtime only, which
# gives full, trustworthy coverage for INSTALLS in the reporting week. It does NOT
# cover the other 47 weeks of the Pace baseline (never pulled) or the matching
# eval population (eval uploads in this file belong to evals performed weeks
# before the install date). Scoping rule: use the real date ONLY for rows the
# pull actually attempted (this week's installs); fall back to the proxy for
# every other row, so Pace/bands/Trends still have a complete 48-week history
# to compute against. Getting this wrong once wiped every Pace figure to "-".
a['idone']=a['Install Packet Sent'].astype(str).str.strip().str.lower().eq('yes')
a['edone']=a['qsd'].notna()
EVAL_REAL_ACTIVE = False
eup = None     # module-level defaults: these stay unset on a proxy-basis run.
               # Without them, any run WITHOUT the eval upload CSV crashed --
               # eup and ulag_eval_real were only ever bound inside the
               # upload-CSV branch, so the plain command path was broken.
a['ulag_eval_real']=float('nan')
# Boots on Ground: receipt of the request to the day the crew is actually on
# site. Eval side has no real RFQ-received field in the export -- Created On
# is already the standing proxy used everywhere else on this report (Method
# & Notes r59). Install side uses the real PO Received Date.
a['cod']=pd.to_datetime(a['Created On'],errors='coerce')
a['prd']=pd.to_datetime(a['PO Received Date'],errors='coerce')
a['ebog']=(a['edt']-a['cod']).dt.total_seconds()/86400
a['ibog']=(a['idt']-a['prd']).dt.total_seconds()/86400
# Holiday-adjusted counts per the framework: "Holidays are skipped EVERYWHERE
# -- the displayed day-count, the on-time window, and the red-zone / flag
# positions all use the same holiday-adjusted count." East eval + all installs
# are calendar mode (holidays only); West eval is the sole business-day section
# (weekends AND holidays). Region decides eval mode, so both are computed and
# the right one is selected per row once region is known.
a['ebog_cal']=day_count_series(a['cod'],a['edt'],'calendar')
a['ebog_bus']=day_count_series(a['cod'],a['edt'],'business')
a['ibog_cal']=day_count_series(a['prd'],a['idt'],'calendar')
# Per-row mode selection: West Eval is the only business-day section
# ("West Eval is the only business-day section", framework Table 0 + source note).
a['ebog_adj']=np.where(a['region']=='West',a['ebog_bus'],a['ebog_cal'])
a['ibog_adj']=a['ibog_cal']
a['ebog_adj']=pd.to_numeric(a['ebog_adj'],errors='coerce')
a['ibog_adj']=pd.to_numeric(a['ibog_adj'],errors='coerce')
# Broader re-eval detection than the upload-lag outlier check -- this looks
# at every currently active project's own note text, not just the ones whose
# lag happened to look wrong this specific week. A project can be a genuine
# re-eval without ever producing a negative-lag artifact (e.g. if it's still
# pending and hasn't uploaded anything yet).
a['_is_reeval']=a['Most Recent Timeline Note'].astype(str).str.upper().str.contains('RE-EVAL|REEVAL',regex=True,na=False)
# "Current open" for the Re-evals tally: still in progress, not sitting in
# Invoicing And Close (the terminal stage -- 18,124 of ~22k rows nationwide).
# Confirmed against real data this maps to a plausible current working book
# size (109-140 for busy locations) rather than the full historical count
# (which ran into the thousands and was the wrong denominator).
_STAGECOL='Active Stage (Active Stage) (Project Management Flow)'
a['_is_open']=a[_STAGECOL]!='Invoicing And Close' if _STAGECOL in a.columns else True
# 'ulag' is the EVAL-side lag (Quote Sent - Eval Date) everywhere in the report --
# Locations' Avg Upload column, Upload Visualization, Trends all read it as such.
# It must NOT be redefined to install lag here: that overwrite is what previously
# collapsed Avg Upload coverage from all 52 locations down to 5.
if USE_REAL_UPLOAD and upl is not None:
    _pulled = a['_pid'].isin(upl.index)
    a.loc[_pulled,'idone'] = a.loc[_pulled,'iup_local'].notna()
    # real INSTALL lag lives in its own column, scoped to only the pulled rows,
    # so it never collides with the eval-side 'ulag' used elsewhere
    a['ulag_install_real']=(a['iup_local']-a['idt']).dt.total_seconds()/86400
    EVAL_UPLOAD_CSV = os.environ.get('AA_EVAL_UPLOAD_CSV','')
    eup=None
    if EVAL_UPLOAD_CSV and os.path.exists(EVAL_UPLOAD_CSV):
        eup = pd.read_csv(EVAL_UPLOAD_CSV, dtype=str).set_index('project')
        eup['eval_upload']=pd.to_datetime(eup['eval_upload'],errors='coerce',utc=True)
        a['eu2_raw']=a['_pid'].map(eup['eval_upload'])
        a['eup_local2']=[local_naive(t,s) for t,s in zip(a['eu2_raw'],a['_ustate'])]
        a['ulag_eval_real']=(a['eup_local2']-a['edt']).dt.total_seconds()/86400
        # This line was missing entirely: eup was loaded and used for the lag
        # column, but eval COMPLETION kept reading the Quote-Sent proxy no matter
        # what CSV was provided. Every eval count on every tab was silently still
        # proxy-based even with AA_EVAL_UPLOAD_CSV supplied -- this is the actual
        # completion switch, mirroring the install-side line above.
        _pulled_eval = a['_pid'].isin(eup.index)
        a.loc[_pulled_eval,'edone'] = a.loc[_pulled_eval,'eup_local2'].notna()
        # cevid above only ever read the install-side CSV -- the 30 eval-side
        # cancellation-evidence flags never reached the Scoreboard at all, not
        # even mislabeled. Merging both sides in with OR so nothing is invisible.
        if 'cancellation_evidence' in eup.columns:
            eup['cancellation_evidence']=eup['cancellation_evidence'].astype(str).str.strip().str.lower().eq('true')
            _eval_cevid = a['_pid'].map(eup['cancellation_evidence']).fillna(False)
            a['cevid'] = a['cevid'] | _eval_cevid
    else:
        a['ulag_eval_real']=float('nan')
    EVAL_REAL_ACTIVE = eup is not None
    DOCHEALTH_CSV = os.environ.get('AA_DOCHEALTH_CSV','')
    if DOCHEALTH_CSV and os.path.exists(DOCHEALTH_CSV):
        dh = pd.read_csv(DOCHEALTH_CSV, dtype=str).set_index('project')
        dh['unclassified']=pd.to_numeric(dh['unclassified'],errors='coerce')
        dh['excluded']=pd.to_numeric(dh['excluded'],errors='coerce')
        a['unclassified']=a['_pid'].map(dh['unclassified'])
        a['excluded']=a['_pid'].map(dh['excluded'])
    else:
        a['unclassified']=pd.NA; a['excluded']=pd.NA
a['lf']=pd.to_numeric(a['Linear Feet (CadQuoting) (CAD Quoting)'],errors='coerce')
a['ec']=pd.to_numeric(a['Number Of Call Attempts Eval'],errors='coerce')
a['ic']=pd.to_numeric(a['Number Of Call Attempts Install'],errors='coerce')
a['ulag']=(a['qsd']-a['edt']).dt.total_seconds()/86400   # eval-side proxy, always

EW=a[a['ewk']==WEEK].copy(); IW=a[a['iwk']==WEEK].copy()

# ---- Pace baseline: Sep 2025 through the reporting week ----
BASE_FROM=pd.Timestamp(os.environ.get('AA_BASE_FROM','2025-09-01'))
BWEEKS=pd.date_range(BASE_FROM,WEEK,freq='W-MON')
TIER_A_WEEKS,TIER_A_MEAN,TIER_B_WEEKS=12,0.92,6

def series_for(loc,metric):
    wkcol,tcol,done=('ewk','Eval Technician','edone') if metric=='Eval' else ('iwk','Install Technician','idone')
    sub=a[a[tcol].astype(str).str.strip()==loc]
    cnt=sub[sub[done]].groupby(wkcol).size()
    hit=[w for w in BWEEKS if w in cnt.index]
    if not hit: return None
    first=min(hit)
    return pd.Series([int(cnt.get(w,0)) for w in BWEEKS if w>=first],
                     index=[w for w in BWEEKS if w>=first])

def pace_stats(loc,metric):
    sr=series_for(loc,metric)
    if sr is None or len(sr)<2: return None
    n=len(sr); mean=float(sr.mean()); sd=float(sr.std(ddof=1))
    tier='A' if (n>=TIER_A_WEEKS and mean>=TIER_A_MEAN) else ('B' if n>=TIER_B_WEEKS else 'C')
    return dict(n=n,mean=mean,sd=sd,tier=tier,series=sr,
                lo2=max(0,mean-2*sd),lo1=max(0,mean-sd),hi1=mean+sd,hi2=mean+2*sd,
                cons=(sd/mean**0.5) if mean>0 else None)

def zone_of(v,st):
    if st is None or st['tier']=='C': return None,None
    if v> st['hi2']: return '\u25b2\u25b2 Well ahead',GRN2_Z
    if v> st['hi1']: return '\u25b2 Ahead',GRN_Z
    if v< st['lo2']: return '\u25bc\u25bc ALERT',RED_Z
    if v< st['lo1']: return '\u25bc Caution',YEL_Z
    return '\u25cf On pace',NEU_Z

# directional zone fills — upward is good
GRN2_Z='A8D8A0'; GRN_Z='D9EAD3'; NEU_Z='F2F2F2'; YEL_Z='FFE699'; RED_Z='F4B7BE'

WEEKS=[WEEK-pd.Timedelta(weeks=i) for i in range(WOW_N-1,-1,-1)]
REGIONS=['East','Central','West']

a['itos']=pd.to_numeric(a['Install Time On Site (min)'],errors='coerce')
a['etos']=pd.to_numeric(a['Eval Time On Site (min)'],errors='coerce')
TOS_LO,TOS_HI,TOS_MIN_N,TOS_MIN_COV=15,720,20,0.20
def tos_for(loc,metric='Install'):
    if metric=='Install':
        sub=a[(a['Install Technician'].astype(str).str.strip()==loc)&(a['iwk'].isin(BWEEKS))]
        col='itos'
    else:
        sub=a[(a['Eval Technician'].astype(str).str.strip()==loc)&(a['ewk'].isin(BWEEKS))]
        col='etos'
    if not len(sub): return 'n/a'
    raw=sub[col].dropna(); raw=raw[raw>0]
    cl=raw[(raw>=TOS_LO)&(raw<=TOS_HI)]
    if len(cl)>=TOS_MIN_N and len(raw)/len(sub)>=TOS_MIN_COV: return round(float(cl.mean()))
    return 'insufficient'

wb=Workbook(); del wb['Sheet']

# ---------------- shared iconography ----------------
# One vocabulary used on every tab, so a title can be short and the key explains it.
IC_EVAL='\u25cf'; IC_INST='\u25cf'                 # filled dot, coloured red / blue
IC_DONE='\u2713'; IC_PEND='\u25cb'                 # tick / hollow circle
IC_WELL='\u25b2\u25b2'; IC_AHEAD='\u25b2'
IC_ONPACE='\u25cf'; IC_CAUTION='\u25bc'; IC_ALERT='\u25bc\u25bc'
IC_UP='\u2191'; IC_FLAT='\u2192'; IC_DOWN='\u2193'
IC_SC='\u25a0'; IC_DEALER='\u25a1'; IC_NONE='\u2014'

# ---------------- measure tints ----------------
# One light tint per KIND of number, used wherever that measure appears on any
# tab. Deliberately pale: status fills (zone, rate, pending) sit on top and win.
T_EVAL  ='FDEEF0'   # counting evaluations
T_INST  ='E8F1F9'   # counting installations
T_LF    ='FDF3E3'   # lineal feet
T_CALLS ='EFEAF6'   # call attempts
T_UPLOAD='E4F2F1'   # upload / turnaround days
T_TOS   ='EDF2E9'   # time on site

CHART_TINT={'eval':'F6C9CE','install':'BFDCF0','lf':'E7D9F0',
            'calls':'D9CCEA','upload':'B9E0DE','tos':'CFE0C4'}

MEASURE_TINT={
 'Evals \u2713':T_EVAL,'Evals pend':T_EVAL,'Eval replacement':T_EVAL,
 'Evals completed':T_EVAL,'RFQs received':T_EVAL,'Eval replacement %':T_EVAL,
 'Installs \u2713':T_INST,'Installs pend':T_INST,'Install Pace/wk':T_INST,
 'This week vs Pace':T_INST,'Install replacement':T_INST,'Installs completed':T_INST,
 'POs received':T_INST,'Install replacement %':T_INST,
 'Total LF':T_LF,'Lineal feet installed':T_LF,
 'Avg Calls':T_CALLS,'Avg call attempts':T_CALLS,
 'Eval quote-sent (d)':T_UPLOAD,
 'Eval time on site':T_TOS,'Install time on site':T_TOS,
}

def mtint(ws,row,col,header):
    """Apply the measure tint for this column, unless a status fill already owns it."""
    t=MEASURE_TINT.get(header)
    if not t: return
    cel=ws.cell(row=row,column=col)
    rgb=cel.fill.fgColor.rgb if cel.fill and cel.fill.fgColor else None
    if rgb in (None,'00000000'): cel.fill=P(t)

ZONE_ICON={'Well ahead':IC_WELL,'Ahead':IC_AHEAD,'On pace':IC_ONPACE,
           'Caution':IC_CAUTION,'ALERT':IC_ALERT}

def style_line(lc,specs,title,ytitle):
    """East Region house style: compact, diamond on primary, circles on the rest."""
    lc.style=2; lc.height=9; lc.width=17; lc.title=title
    for k,srs in enumerate(lc.series):
        col,dash,mark,wid=specs[k]
        srs.graphicalProperties=GraphicalProperties()
        lp=LineProperties(solidFill=col,w=wid)
        if dash: lp.prstDash=dash
        srs.graphicalProperties.line=lp
        srs.marker=Marker(symbol=mark,size=5) if mark else Marker(symbol='none')
        srs.smooth=False
    lc.legend.position='r'          # never 'b': it lands on the date labels
    lc.x_axis.delete=False; lc.y_axis.delete=False
    lc.y_axis.majorGridlines=ChartLines()
    lc.x_axis.title='Week (Monday)'; lc.y_axis.title=ytitle

def key_block(ws,row,col,items,title='KEY'):
    """Boxed legend parked clear of the table, so it never reads as data."""
    med=Side(style='medium',color=AA_BLUE)
    h=ws.cell(row=row,column=col,value=title)
    h.font=Font(name=FN,size=9,bold=True,color='FFFFFF'); h.fill=FBLUE
    h.alignment=Alignment(horizontal='center',vertical='center')
    h.border=Border(left=med,right=med,top=med,bottom=thin)
    for i,(icon,label,fill,colr) in enumerate(items,1):
        cel=ws.cell(row=row+i,column=col,value=f'{icon}   {label}')
        cel.font=Font(name=FN,size=9,bold=True,color=colr or '000000')
        if fill: cel.fill=P(fill)
        cel.alignment=Alignment(horizontal='left',vertical='center',indent=1)
        cel.border=Border(left=med,right=med,top=thin,
                          bottom=med if i==len(items) else thin)
        ws.row_dimensions[row+i].height=14
    ws.column_dimensions[L(col)].width=24
    return row+len(items)+1

ZONE_KEY=[(IC_WELL,'Well ahead',GRN2_Z,None),(IC_AHEAD,'Ahead',GRN_Z,None),
          (IC_ONPACE,'On pace',NEU_Z,None),(IC_CAUTION,'Caution',YEL_Z,None),
          (IC_ALERT,'Alert',RED_Z,None)]

def vertical_labels(axis, size=800):
    """Tilt category tick labels 45 degrees, under the bar, standard Excel style.
    Forces axPos to the bottom: without it Excel will not place category labels
    at all on some bar/line combinations. tickLblSkip=1 forces every label to
    draw -- without it Excel can decide on its own that labels won't fit and
    silently skip most of them, which is the most likely reason names have
    vanished on some builds even though the underlying data was always correct."""
    axis.axPos = 'b'
    axis.tickLblSkip = 1
    axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=-2700000, vert='horz', wrap='square',
                                  anchor='ctr', anchorCtr='1'),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size)),
                     endParaRPr=CharacterProperties(sz=size))])

def titles(ws,t,s,big=False):
    c1=ws.cell(row=1,column=1,value=t); c1.font=TITLE16 if big else TITLE15
    c1.alignment=Alignment(horizontal='center',vertical='center')
    c2=ws.cell(row=2,column=1,value=s); c2.font=SUBT
    c2.alignment=Alignment(horizontal='center',vertical='center')

def hdr(ws,row,labels,widths=None):
    for i,t in enumerate(labels,1):
        c=ws.cell(row=row,column=i,value=t)
        c.font=HDW; c.fill=FBLUE; c.alignment=CTR; c.border=BOX
    if widths:
        for i,w in enumerate(widths,1): ws.column_dimensions[L(i)].width=w

def band(ws,row,text,span):
    for cc in range(1,span+1):
        c=ws.cell(row=row,column=cc); c.fill=FBLUE; c.border=BOX
    c=ws.cell(row=row,column=1,value=text); c.font=HDW; c.alignment=LFT

def caption(ws,row,text,span=10):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    c=ws.cell(row=row,column=1,value='▸  '+text); c.font=NOTE
    c.alignment=Alignment(vertical='center',wrap_text=True)
    ws.row_dimensions[row].height=26

def kpi_fill(p): return P(KPI_G) if p>=0.90 else (P(KPI_A) if p>=0.80 else P(KPI_R))

def stat(W,done,tcol=None,loc=None,reg=None):
    s=W
    if loc is not None: s=s[s[tcol].astype(str).str.strip()==loc]
    if reg is not None: s=s[s['region']==reg]
    ok=int(s[done].sum())
    return ok, len(s)-ok, s

# ==================================================== SCOREBOARD
ws=wb.create_sheet('Scoreboard')
for col,w in zip('ABCDEFGHI',[40,13,13,13,15,15,15,15,15]): ws.column_dimensions[col].width=w
ws.column_dimensions['D'].width=42
ws.column_dimensions['E'].width=46
ws.row_dimensions[1].height=40
if os.path.exists(LOGO):
    from openpyxl.drawing.image import Image as XLImage
    img=XLImage(LOGO); ratio=img.width/max(img.height,1)
    img.height=44; img.width=int(44*ratio); img.anchor='A1'
    ws.add_image(img)
_t3=ws.cell(row=3,column=1,value='AA Weekly Report'); _t3.font=TITLE16
_t3.alignment=Alignment(horizontal='center',vertical='center')
_basis = (f"installs: REAL upload dates / evals: {'REAL upload dates' if EVAL_REAL_ACTIVE else 'Quote Sent proxy'}"
          if (USE_REAL_UPLOAD and upl is not None) else 'Quote Sent / Packet Sent proxy')
_t4=ws.cell(row=4,column=1,value=(f'Week of {WEEK.strftime("%b %-d")} \u2013 {(WEEK+pd.Timedelta(days=6)).strftime("%b %-d, %Y")}'
          f'   \u00b7   nationwide   \u00b7   cancelled projects excluded   \u00b7   completion basis: {_basis}'))
_t4.font=SUBT; _t4.alignment=Alignment(horizontal='center',vertical='center')

PREV=WEEK-pd.Timedelta(weeks=1)
EP=a[a['ewk']==PREV]; IP=a[a['iwk']==PREV]
ec,ep,_=stat(EW,'edone'); ic,ip,_=stat(IW,'idone')
ecp=int(EP['edone'].sum()); icp=int(IP['idone'].sum())
lf=IW['lf']; ulag=EW['ulag'].dropna()
_ins_lag_real = a.loc[a['_pid'].isin(upl.index),'ulag_install_real'].dropna() if (USE_REAL_UPLOAD and upl is not None) else pd.Series(dtype=float)
rfq=int((a['cro']==WEEK).sum()) if 'cro' in a else int((pd.to_datetime(a['Created On'],errors='coerce').apply(wk_of)==WEEK).sum())
po =int((pd.to_datetime(a['PO Received Date'],errors='coerce').apply(wk_of)==WEEK).sum())
ehist=pd.Series([int(a[a['ewk']==w]['edone'].sum()) for w in WEEKS])
ihist=pd.Series([int(a[a['iwk']==w]['idone'].sum()) for w in WEEKS])
emu,esd=float(ehist.mean()),float(ehist.std(ddof=1))
imu,isd=float(ihist.mean()),float(ihist.std(ddof=1))

def verdict(v,mu,sd):
    if v>mu+2*sd: return '\u25b2\u25b2 Well above normal',GRN2_Z
    if v>mu+sd:   return '\u25b2 Above normal',GRN_Z
    if v<mu-2*sd: return '\u25bc\u25bc WELL BELOW normal',RED_Z
    if v<mu-sd:   return '\u25bc Below normal',YEL_Z
    return '\u25cf Normal',NEU_Z

# ---------------- 1. headline ----------------
r=6
sect_row=r
for cc in range(1,10): ws.cell(row=r,column=cc).fill=FBLUE
c=ws.cell(row=r,column=1,value='THE WEEK AT A GLANCE'); c.font=HDW; c.alignment=LFT
r+=1
key_block(ws,6,11,ZONE_KEY+[(IC_UP,'queue growing',KPI_A,None),
                            (IC_FLAT,'queue level',NEU_Z,None),
                            (IC_DOWN,'queue draining',GRN_Z,None),
                            (IC_SC,'Service Center',CHAN_GREEN,None),
                            (IC_DEALER,'Dealer',CHAN_AMBER,None)],'KEY')
hdr(ws,r,['','Completed','Pending','Completion rate','vs last week','Pace/wk','','Verdict','Locations'])
r+=1
for lbl,done_n,pend_n,tot_n,prev_n,mu,sd,nloc in [
        ('Evaluations',ec,ep,len(EW),ecp,emu,esd,EW['Eval Technician'].nunique()),
        ('Installations',ic,ip,len(IW),icp,imu,isd,IW['Install Technician'].nunique())]:
    ws.cell(row=r,column=1,value=lbl).font=BD
    ws.cell(row=r,column=3).comment=None
    c=ws.cell(row=r,column=2,value=done_n); c.font=Font(name=FN,size=20,bold=True,color=AA_BLUE); c.alignment=CTR
    ws.cell(row=r,column=3,value=pend_n).alignment=CTR
    rc=ws.cell(row=r,column=4,value=done_n/tot_n if tot_n else 0)
    rc.number_format='0%'; rc.alignment=CTR; rc.fill=kpi_fill(done_n/max(tot_n,1)); rc.font=BD
    d=done_n-prev_n
    dc=ws.cell(row=r,column=5,value=d); dc.number_format='+0;-0;0'; dc.alignment=CTR; dc.font=BD
    dc.fill=P(GRN_Z) if d>0 else (P(RED_Z) if d<0 else P(NEU_Z))
    pc=ws.cell(row=r,column=6,value=round(mu,1)); pc.number_format='0.0'; pc.alignment=CTR
    vtxt,vfill=verdict(done_n,mu,sd)
    vc=ws.cell(row=r,column=8,value=vtxt); vc.fill=P(vfill); vc.font=BD; vc.alignment=CTR
    ws.cell(row=r,column=9,value=int(nloc)).alignment=CTR
    for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
    ws.row_dimensions[r].height=28
    r+=1
    # Normal range: kept as data, not a visible column -- collapsed into a
    # hidden row directly beneath its summary row rather than a seventh
    # column, per request. Same 1-2-sigma bands the verdict already uses.
    ws.cell(row=r,column=1,value=f'      Normal range ({lbl.lower()})').font=NOTE
    ws.cell(row=r,column=7,value=f'{max(0,mu-2*sd):.0f}\u2013{mu+2*sd:.0f}').alignment=CTR
    ws.row_dimensions[r].hidden=True
    r+=1
ws.cell(row=r,column=1,value='Lineal feet installed').font=BD
c=ws.cell(row=r,column=2,value=int(lf.sum())); c.number_format='#,##0'; c.font=BD; c.alignment=CTR
ws.cell(row=r,column=3,value=f'{int(lf.notna().sum())} of {len(IW)} entered').font=NOTE
ws.cell(row=r,column=6,value=round(float(ulag.mean()),1) if len(ulag) else '\u2014').alignment=CTR
ws.cell(row=r,column=5,value='Eval quote-sent lag (d, proxy):').font=NOTE
for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
r+=1
ws.cell(row=r,column=1,value='Install upload lag (d, real)').font=NOTE
c=ws.cell(row=r,column=6,value=round(float(_ins_lag_real.mean()),1) if len(_ins_lag_real) else 'not pulled this run')
c.alignment=CTR
for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
r+=2

# ---------------- 2. replacement rate ----------------
for cc in range(1,10): ws.cell(row=r,column=cc).fill=FBLUE
c=ws.cell(row=r,column=1,value='KEEP-UP RATE — can we complete what comes in?'); c.font=HDW; c.alignment=LFT
r+=1
hdr(ws,r,['','Received this week','Completed this week','Keep-up rate','Queue direction',
          'Net change','','',''])
r+=1
for lbl,inn,outn in [('RFQs \u2192 Evaluations',rfq,ec),('POs \u2192 Installations',po,ic)]:
    ws.cell(row=r,column=1,value=lbl).font=BD
    c=ws.cell(row=r,column=2,value=inn); c.font=Font(name=FN,size=14,bold=True); c.alignment=CTR
    ws.cell(row=r,column=3,value=outn).alignment=CTR
    pctv=(outn/inn) if inn else 0
    rc=ws.cell(row=r,column=4,value=pctv); rc.number_format='0%'; rc.alignment=CTR; rc.font=BD
    dirn='\u2193 FALLING BEHIND' if pctv<0.85 else ('\u2192 tight' if pctv<0.95 else ('\u2191 keeping up' if pctv<1.05 else '\u2191 catching up'))
    rc.fill=P(KPI_R) if pctv<0.85 else (P(KPI_A) if pctv<0.95 else P(GRN_Z))
    dc=ws.cell(row=r,column=5,value=dirn); dc.alignment=CTR; dc.font=BD
    dc.fill=P(KPI_R) if 'FALLING' in dirn else (P(KPI_A) if 'tight' in dirn else P(GRN_Z))
    nc=ws.cell(row=r,column=6,value=inn-outn); nc.number_format='+0;-0;0'; nc.alignment=CTR
    for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
    r+=1
caption(ws,r,'Keep-up rate = work COMPLETED \u00f7 work RECEIVED, same week. 100% means we finished '
        'exactly what came in. Below 100% we fell behind and the queue GREW by the Net change shown; '
        'above 100% we completed more than arrived and the queue drained. Five-week history on the '
        'Trends tab.',9)
r+=2

# ---------------- 3. by region ----------------
for cc in range(1,10): ws.cell(row=r,column=cc).fill=FBLUE
c=ws.cell(row=r,column=1,value='BY REGION'); c.font=HDW; c.alignment=LFT
r+=1
hdr(ws,r,['Region','Evals \u2713','Evals pending','Eval completion %','Installs \u2713','Installs pending',
          'Install completion %','Eval keep-up %','Install keep-up %'])
r+=1
RSTART=r
for reg in REGIONS:
    e_ok,e_pd,e=stat(EW,'edone',reg=reg); i_ok,i_pd,i=stat(IW,'idone',reg=reg)
    ws.cell(row=r,column=1,value=reg).font=BD
    for cc,v in ((2,e_ok),(3,e_pd),(5,i_ok),(6,i_pd)):
        cel=ws.cell(row=r,column=cc,value=v); cel.alignment=CTR
    for cc,ok,tot_ in ((4,e_ok,len(e)),(7,i_ok,len(i))):
        pp=ok/tot_ if tot_ else 0
        cel=ws.cell(row=r,column=cc,value=pp); cel.number_format='0%'; cel.alignment=CTR; cel.fill=kpi_fill(pp)
    rfq_r=int(((pd.to_datetime(a['Created On'],errors='coerce').apply(wk_of)==WEEK)&(a['region']==reg)).sum())
    po_r =int(((pd.to_datetime(a['PO Received Date'],errors='coerce').apply(wk_of)==WEEK)&(a['region']==reg)).sum())
    for cc,inn,outn in ((8,rfq_r,e_ok),(9,po_r,i_ok)):
        pv=(outn/inn) if inn else 0
        cel=ws.cell(row=r,column=cc,value=pv); cel.number_format='0%'; cel.alignment=CTR; cel.font=BD
        cel.fill=P(KPI_R) if pv<0.85 else (P(KPI_A) if pv<0.95 else P(GRN_Z))
    for cc in (3,6):
        if ws.cell(row=r,column=cc).value: ws.cell(row=r,column=cc).fill=P(KPI_A)
    for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
    r+=1
REND=r-1
ws.conditional_formatting.add(f'B{RSTART}:B{REND}',DataBarRule(start_type='min',end_type='max',color=AA_RED,showValue=True))
ws.conditional_formatting.add(f'E{RSTART}:E{REND}',DataBarRule(start_type='min',end_type='max',color=AA_BLUE,showValue=True))


# ---- region share of national volume, donut ----
# Fixed layout: table starts L6, chart anchored ~L12.
_rc=12
ws.cell(row=6,column=_rc,value='REGION SHARE OF VOLUME').font=HDW
for _cc in range(_rc,_rc+2): ws.cell(row=6,column=_cc).fill=FBLUE
ws.cell(row=7,column=_rc,value='Region').font=BD
ws.cell(row=7,column=_rc+1,value='Total completions').font=BD
for _i,reg in enumerate(REGIONS):
    e_ok,_,_=stat(EW,'edone',reg=reg); i_ok,_,_=stat(IW,'idone',reg=reg)
    ws.cell(row=8+_i,column=_rc,value=reg)
    ws.cell(row=8+_i,column=_rc+1,value=e_ok+i_ok)
    for _cc in (_rc,_rc+1): ws.cell(row=8+_i,column=_cc).border=BOX
_don=DoughnutChart(); _don.height=9.5; _don.width=12
_don.title='Region share of national volume'
_don.add_data(Reference(ws,min_col=_rc+1,min_row=7,max_row=7+len(REGIONS)),titles_from_data=True)
_don.set_categories(Reference(ws,min_col=_rc,min_row=8,max_row=7+len(REGIONS)))
from openpyxl.chart.marker import DataPoint as _DP
_REG_COLORS=[soften(AA_BLUE),soften(AA_RED),soften(AA_SILVER)]
_don.series[0].data_points=[_DP(idx=_i,spPr=GraphicalProperties(solidFill=c))
                            for _i,c in enumerate(_REG_COLORS)]
_don.series[0].dLbls=DataLabelList(showPercent=True,showVal=False,showCatName=True,
    showSerName=False,showLegendKey=False,dLblPos='outEnd')
_don.legend=None
_don.anchor=TwoCellAnchor(editAs='twoCell',
    _from=AnchorMarker(col=_rc-1,colOff=0,row=11,rowOff=0),
    to=AnchorMarker(col=_rc+6,colOff=0,row=28,rowOff=0))
ws.add_chart(_don)

# ---- cancellation reasons, THIS REPORTING WEEK ONLY ----
# No "Cancelled On" date exists in the export -- only the code and reason text.
# Scoped to projects whose EVAL or INSTALL date falls in the reporting week,
# same convention as everything else on this tab. This previously pulled the
# entire export history (5,558+ rows) -- fixed to match the rest of the tab.
_ewk_c=pd.to_datetime(df['Eval Date and Time'],errors='coerce').apply(wk_of)
_iwk_c=pd.to_datetime(df['Install Date and Time'],errors='coerce').apply(wk_of)
_wk_mask=(_ewk_c==WEEK)|(_iwk_c==WEEK)
_cx=df.loc[df['Cancellation Code'].notna()&_wk_mask,'Cancellation Code'].astype(str).str.split('-').str[0].str.strip()
_cx=_cx[_cx.str.len()>0].value_counts()
_CANCEL_PALETTE=[soften(c) for c in ['D2001A','004890','A6A6A6','6B7280','FFC000','70AD47','7030A0','C55A11']]
_pc=_rc; _prow=30
ws.cell(row=_prow,column=_pc,value=f'CANCELLATION REASONS \u2014 week of {WEEK.strftime("%b %-d")}').font=HDW
for _cc in range(_pc,_pc+2): ws.cell(row=_prow,column=_cc).fill=FBLUE
ws.cell(row=_prow+1,column=_pc,value='Cancellation reason').font=BD
ws.cell(row=_prow+1,column=_pc+1,value='Count').font=BD
if len(_cx):
    for _i,(reason,cnt) in enumerate(_cx.head(8).items()):
        _rr=_prow+2+_i
        _clr=_CANCEL_PALETTE[_i%len(_CANCEL_PALETTE)]
        rc_=ws.cell(row=_rr,column=_pc,value=reason); rc_.fill=P(_clr)
        rc_.font=Font(name=FN,size=11,bold=True,color='333333')
        cc_=ws.cell(row=_rr,column=_pc+1,value=int(cnt)); cc_.fill=P(_clr); cc_.font=Font(name=FN,size=11,color='333333')
        for _c2 in (_pc,_pc+1): ws.cell(row=_rr,column=_c2).border=BOX
    _pie=PieChart(); _pie.height=9.5; _pie.width=12
    _pie.title=f'Cancellation reasons \u2014 {int(_cx.sum())} this week'
    _pie.add_data(Reference(ws,min_col=_pc+1,min_row=_prow+1,max_row=_prow+1+min(len(_cx),8)),titles_from_data=True)
    _pie.set_categories(Reference(ws,min_col=_pc,min_row=_prow+2,max_row=_prow+1+min(len(_cx),8)))
    _pie.series[0].data_points=[_DP(idx=_i,spPr=GraphicalProperties(
        solidFill=_CANCEL_PALETTE[_i%len(_CANCEL_PALETTE)])) for _i in range(min(len(_cx),8))]
    _pie.series[0].dLbls=DataLabelList(showPercent=True,showVal=False,showCatName=True,
        showSerName=False,showLegendKey=False,dLblPos='bestFit')
    _pie.legend=None
    _pie.anchor=TwoCellAnchor(editAs='twoCell',
        _from=AnchorMarker(col=_pc-1,colOff=0,row=41,rowOff=0),
        to=AnchorMarker(col=_pc+6,colOff=0,row=58,rowOff=0))
    ws.add_chart(_pie)
else:
    caption(ws,_prow+2,"No cancellations recorded against this week's eval/install dates.",2)
for _cc in range(_rc,_rc+8): ws.column_dimensions[L(_cc)].hidden=False
r+=1

# ---------------- 4. channel ----------------
for cc in range(1,10): ws.cell(row=r,column=cc).fill=FBLUE
c=ws.cell(row=r,column=1,value='SERVICE CENTERS vs DEALERS'); c.font=HDW; c.alignment=LFT
r+=1
hdr(ws,r,['Channel','Evals \u2713','Installs \u2713','Share of all work','Incomplete','Lineal feet','Locations','',''])
r+=1
for lbl,pred in [('Service Centers',True),('Dealers',False)]:
    e=EW[EW['Eval Technician'].apply(is_sc)==pred]; i=IW[IW['Install Technician'].apply(is_sc)==pred]
    eo=int(e['edone'].sum()); io_=int(i['idone'].sum())
    outst=int((~e['edone']).sum())+int((~i['idone']).sum())
    ws.cell(row=r,column=1,value=lbl).font=BD
    ws.cell(row=r,column=2,value=eo).alignment=CTR
    ws.cell(row=r,column=3,value=io_).alignment=CTR
    shr=(eo+io_)/max(ec+ic,1)
    sc_=ws.cell(row=r,column=4,value=shr); sc_.number_format='0%'; sc_.alignment=CTR; sc_.font=BD
    oc=ws.cell(row=r,column=5,value=outst); oc.alignment=CTR
    if outst: oc.fill=P(KPI_A)
    cel=ws.cell(row=r,column=6,value=int(i['lf'].sum())); cel.number_format='#,##0'; cel.alignment=CTR
    ws.cell(row=r,column=7,value=int(pd.concat([e['Eval Technician'],i['Install Technician']]).nunique())).alignment=CTR
    ws.cell(row=r,column=1).fill=P(CHAN_GREEN if pred else CHAN_AMBER)
    for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
    r+=1
caption(ws,r,"Share of all work = that channel's completions as a percentage of the national total. "
        "Incomplete = worked this week but the next step is not yet recorded. Completion rates are on "
        "the headline and By Region blocks above.",9)
r+=1

# ---------------- 5. movers ----------------
for cc in range(1,10): ws.cell(row=r,column=cc).fill=FBLUE
c=ws.cell(row=r,column=1,value='HIGHEST AND LOWEST VOLUME \u2014 this week'); c.font=HDW; c.alignment=LFT
r+=1
hdr(ws,r,['Location','Channel','Evals \u2713','Installs \u2713','Total completions',
          'Share of national','vs last week','',''])
r+=1
vol=[]
for loc in set(pd.concat([EW['Eval Technician'],IW['Install Technician']]).dropna().astype(str).str.strip()):
    en=int(EW[EW['Eval Technician'].astype(str).str.strip()==loc]['edone'].sum())
    inn=int(IW[IW['Install Technician'].astype(str).str.strip()==loc]['idone'].sum())
    was=int(EP[EP['Eval Technician'].astype(str).str.strip()==loc]['edone'].sum())+\
        int(IP[IP['Install Technician'].astype(str).str.strip()==loc]['idone'].sum())
    if en+inn: vol.append((en+inn,loc,en,inn,was))
vol.sort(reverse=True)
NAT=max(ec+ic,1)

def vol_row(rw,tot,loc,en,inn,was,hi):
    ws.cell(row=rw,column=1,value=loc).font=BD
    cc2=ws.cell(row=rw,column=2,value='\u25a0 Service Center' if is_sc(loc) else '\u25a1 Dealer')
    cc2.alignment=CTR; cc2.fill=P(CHAN_GREEN if is_sc(loc) else CHAN_AMBER)
    for cc,v in ((3,en),(4,inn)): ws.cell(row=rw,column=cc,value=v).alignment=CTR
    tc=ws.cell(row=rw,column=5,value=tot); tc.font=BD; tc.alignment=CTR
    tc.fill=P(GRN_Z if hi else NEU_Z)
    sc_=ws.cell(row=rw,column=6,value=tot/NAT); sc_.number_format='0.0%'; sc_.alignment=CTR
    d=tot-was
    dc=ws.cell(row=rw,column=7,value=d); dc.number_format='+0;-0;0'; dc.alignment=CTR
    dc.fill=P(GRN_Z) if d>0 else (P(RED_Z) if d<0 else P(NEU_Z))
    for cc in range(1,10): ws.cell(row=rw,column=cc).border=BOX

for label,rows_,hi in [('HIGHEST VOLUME',vol[:8],True),('LOWEST VOLUME',vol[-8:],False)]:
    lc_=ws.cell(row=r,column=1,value=label); lc_.font=BD; lc_.fill=P(LBLU3)
    for cc in range(1,10): ws.cell(row=r,column=cc).fill=P(LBLU3)
    r+=1
    for tot,loc,en,inn,was in rows_:
        vol_row(r,tot,loc,en,inn,was,hi); r+=1
caption(ws,r,'Ranked by total completions this week, not by change. Share of national shows how much of '
        'the week each location carried. The last column is week-over-week movement, for context only.',9)
r+=2

# ---------------- 6. attention ----------------
pend_rows=[]
for W_,dn,tcol,dcol,met in [(EW,'edone','Eval Technician','edt','Eval'),
                            (IW,'idone','Install Technician','idt','Install')]:
    for _,row_ in W_[~W_[dn]].iterrows():
        pend_rows.append((row_[tcol],row_['region'],met,row_['Project ID'],
                          row_.get('End User'),row_[dcol],row_.get('ProjectCoordinator')))
for cc in range(1,10): ws.cell(row=r,column=cc).fill=FBLUE

c=ws.cell(row=r,column=1,value=f'NEEDS ATTENTION \u2014 {len(pend_rows)} worked but not completed'); c.font=HDW; c.alignment=LFT
r+=1
by_loc={}
for t in pend_rows: by_loc.setdefault(t[0],[]).append(t)
hdr(ws,r,['Location','Region','Incomplete','Why it is incomplete','Projects','','','',''])
r+=1
for loc,items in sorted(by_loc.items(),key=lambda t:-len(t[1]))[:10]:
    ev=sum(1 for t in items if t[2]=='Eval'); it_=len(items)-ev
    why=[]
    if ev: why.append(f'{ev} eval'+('s' if ev>1 else '')+' worked, no quote sent')
    if it_: why.append(f'{it_} install'+('s' if it_>1 else '')+' worked, packet not sent')
    ws.cell(row=r,column=1,value=loc).font=BD
    ws.cell(row=r,column=2,value=items[0][1]).alignment=CTR
    cel=ws.cell(row=r,column=3,value=len(items)); cel.alignment=CTR; cel.font=BD
    cel.fill=P(KPI_A if len(items)<5 else KPI_R)
    ws.cell(row=r,column=4,value='; '.join(why)).alignment=LFT
    ids=', '.join(str(t[3]) for t in items[:6])+(' \u2026' if len(items)>6 else '')
    ws.cell(row=r,column=5,value=ids).font=SM
    for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
    r+=1
caption(ws,r,'"Incomplete" means the work was performed this week but the next step has not been recorded '
        'in AIMS \u2014 no Quote Sent Date on an eval, or Install Packet Sent still No. Full list with '
        'coordinator and timeline note is on Completions Detail.',9)
r+=2

allloc=set(a[a['ewk'].isin(WEEKS)]['Eval Technician'].dropna().astype(str).str.strip()) | \
       set(a[a['iwk'].isin(WEEKS)]['Install Technician'].dropna().astype(str).str.strip())
ws=wb.create_sheet('5 Week Trends')
titles(ws,'5 Week Trends — week over week',
       f'{WOW_N} weeks to {WEEK.strftime("%b %-d")}, split East / Central / West. '
       'Completions and quality on top, then work coming IN against work going OUT.')
RC=46
for _c in range(RC,RC+WOW_N+1): ws.column_dimensions[L(_c)].hidden=True
key_block(ws,4,2+WOW_N+12,[(IC_EVAL,'Evals',LRED3,AA_RED),(IC_INST,'Installs',LBLU3,AA_BLUE)]+ZONE_KEY,
          'KEY \u2014 measure / zone')
hdr(ws,4,['Region / Location','Measure']+[w.strftime('%-m/%-d') for w in WEEKS],[40,34]+[10]*WOW_N)
ws.freeze_panes='C5'      # week dates and the label columns stay visible when scrolling
METS=['Evals ✓','Installs ✓','Total LF','Eval quote-sent (d)','Avg Calls']
r=5
rowmap={}
REGION_TOP=r
for reg in REGIONS:
    band(ws,r,f'   {reg}',2+WOW_N); r+=1
    for met in METS:
        ws.cell(row=r,column=1,value='').font=RG
        ws.cell(row=r,column=2,value=met).font=BD
        mtint(ws,r,2,met)
        _t=MEASURE_TINT.get(met)
        if _t:
            for _cc in range(2,3+WOW_N): ws.cell(row=r,column=_cc).fill=P(_t)
        for j,w in enumerate(WEEKS,3):
            e=a[(a['ewk']==w)&(a['region']==reg)]; i=a[(a['iwk']==w)&(a['region']==reg)]
            if   met=='Evals ✓':       v=int(e['edone'].sum())
            elif met=='Installs ✓':    v=int(i['idone'].sum())
            elif met=='Total LF':      v=int(i['lf'].sum())
            elif met=='Eval quote-sent (d)':v=round(float(e['ulag'].dropna().mean()),1) if e['ulag'].notna().any() else 0
            else:                      v=round(float(pd.concat([e['ec'],i['ic']]).dropna().mean()),1) if len(e)+len(i) else 0
            c=ws.cell(row=r,column=j,value=v); c.font=RG; c.alignment=CTR; c.border=BOX
        for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
        rowmap.setdefault(met,[]).append(r); r+=1
    r+=1

# ---- share-of-week data bars: bar length = this region's share of the national
# total for that exact week, not that region's own min/max across its 5 weeks.
# A full bar means 100% of that week's national volume came from this region. ----
for met in ('Evals ✓','Installs ✓'):
    _rrows=rowmap[met]                 # [East row, Central row, West row]
    for j in range(3,3+WOW_N):         # each week column
        _vals=[ws.cell(row=rr,column=j).value or 0 for rr in _rrows]
        _total=sum(_vals)
        if _total<=0: continue
        _col=L(j)
        _sqref=' '.join(f'{_col}{rr}' for rr in _rrows)
        ws.conditional_formatting.add(_sqref, DataBarRule(
            start_type='num', start_value=0, end_type='num', end_value=_total,
            color=AA_RED if met=='Evals ✓' else AA_BLUE, showValue=True))

# ---- one chart per MEASURE, all three regions together ----
REG_SPEC=[(AA_BLUE,None,'diamond',28000),      # East
          (AA_RED,None,'circle',25400),        # Central
          (AA_SILVER,None,'circle',25400)]     # West
# 2x2 grid instead of a 4-tall vertical stack: Evals/Installs on top, Total LF/
# Avg Calls directly beside them -- keeps the whole block compact and clear of
# whatever table content sits below, instead of reaching down into it.
_grid=[(0,0),(10,0),(0,13),(10,13)]
for _gi,met in enumerate([m for m in METS if m != 'Eval quote-sent (d)']):      # quote-sent graph removed on request
    rows=rowmap[met]
    mc=LineChart()
    for rw in rows:                            # region rows are not contiguous
        mc.add_data(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=rw,max_row=rw),from_rows=True)
    mc.set_categories(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=4))
    for k,reg in enumerate(REGIONS):
        mc.series[k].tx=SeriesLabel(v=reg)
    style_line(mc,REG_SPEC,f'{met} \u2014 by region','')
    mc.height=7.8; mc.width=17
    _coff,_roff=_grid[_gi]
    mc.anchor=TwoCellAnchor(editAs='oneCell',
        _from=AnchorMarker(col=2+WOW_N+1+_coff,colOff=0,row=REGION_TOP-1+_roff,rowOff=0),
        to=AnchorMarker(col=2+WOW_N+10+_coff,colOff=0,row=REGION_TOP+11+_roff,rowOff=0))
    ws.add_chart(mc)

# ---- nationwide actual vs Pace band (mean +/- 2 sigma) ----
r+=1
band(ws,r,'◆  Nationwide vs Pace band',2+WOW_N); r+=1
caption(ws,r,'The heavy line is what actually happened each week. The flat centre line is the Pace — the '
        'mean weekly output across the full baseline. The two dotted lines are Pace ± 2 sigma. Roughly 19 '
        'weeks in 20 should land between them; a point outside is genuinely unusual, not ordinary variation.',
        2+WOW_N)
r+=1
NATROWS={}
for met,wkcol,done in [('Evals','ewk','edone'),('Installs','iwk','idone')]:
    hist=pd.Series([int(a[a[wkcol]==w][done].sum()) for w in BWEEKS])
    mu=float(hist.mean()); sd=float(hist.std(ddof=1))
    lo=max(0,mu-2*sd); hi=mu+2*sd
    for lbl,vals,fmt in [(f'{met} actual',[int(a[a[wkcol]==w][done].sum()) for w in WEEKS],'0'),
                         (f'{met} Pace (mean)',[round(mu,1)]*WOW_N,'0.0'),
                         (f'{met} +2 sigma',[round(hi,1)]*WOW_N,'0.0'),
                         (f'{met} -2 sigma',[round(lo,1)]*WOW_N,'0.0')]:
        ws.cell(row=r,column=1,value=lbl).font=BD if 'actual' in lbl else RG
        ws.cell(row=r,column=2,value=met).font=RG
        for j,v in enumerate(vals,3):
            cel=ws.cell(row=r,column=j,value=v); cel.font=RG; cel.alignment=CTR
            cel.number_format=fmt; cel.border=BOX
        for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
        NATROWS.setdefault(met,[]).append(r); r+=1
    r+=1

# ---- total RFQs and POs received vs completed ----
a['cro']=pd.to_datetime(a['Created On'],errors='coerce').apply(wk_of)
a['pow']=pd.to_datetime(a['PO Received Date'],errors='coerce').apply(wk_of)
r+=1
band(ws,r,'\u25c6  Intake \u2014 total RFQs and POs received vs completed',2+WOW_N); r+=1
caption(ws,r,'RFQs received is new eval work arriving; POs received is new install work released. '
        'Compare each against what was COMPLETED the same week. Keep-up rate is completed divided '
        'by received: 100% means we finished exactly what arrived that week. Below 100% we fell '
        'behind and the queue grew; above 100% we drained it.',2+WOW_N)
r+=1
INTROWS={}
for metric,incol,inlbl,outlbl,done,wkcol in [
        ('Eval','cro','RFQs received (IN)','Evals completed (OUT)','edone','ewk'),
        ('Install','pow','POs received (IN)','Installs completed (OUT)','idone','iwk')]:
    rowset=[]
    ins=[int((a[incol]==w).sum()) for w in WEEKS]
    outs=[int(a[a[wkcol]==w][done].sum()) for w in WEEKS]
    for lbl,vals,fmt in [(inlbl,ins,'0'),(outlbl,outs,'0'),
                         ('Keep-up rate',
                          [(o/i) if i else None for i,o in zip(ins,outs)],'0%')]:
        ws.cell(row=r,column=1,value=lbl).font=BD if 'Keep-up' not in lbl else RG
        ws.cell(row=r,column=1).alignment=LFT
        ws.cell(row=r,column=2,value=metric).font=RG
        for j,v in enumerate(vals,3):
            cel=ws.cell(row=r,column=j,value=v if v is not None else '\u2014')
            cel.font=RG; cel.alignment=CTR; cel.border=BOX; cel.number_format=fmt
            if 'Keep-up' in lbl and isinstance(v,float):
                cel.fill=P(KPI_R) if v<0.85 else (P(KPI_A) if v<0.95 else P(GRN_Z))
        for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
        rowset.append(r); r+=1
    INTROWS[metric]=rowset
    r+=1

# ---- per-location week over week, collapsible, chart beside the numbers ----
# Chart series live in HIDDEN columns (DC..) so the visible table stays clean.
# TwoCellAnchor + editAs='twoCell' makes each chart collapse with its rows.
ws.sheet_properties.outlinePr.summaryBelow=False
BLOCK_ROWS=22                      # generous: a cramped chart is an unreadable chart                      # vertical space each open location gets
DC=20                              # first hidden data column
for cc in range(DC,DC+WOW_N+2):
    ws.column_dimensions[L(cc)].hidden=True
ZR,ZY,ZN,ZG,ZG2='FCE4E6','FFF2CC','F7F8F9','EAF3E4','D9EAD3'   # light zone fills

r+=1
band(ws,r,'\u25c6  By Location \u2014 week over week   (click + in the left margin to open, \u2212 to close)',
     2+WOW_N)
r+=1
caption(ws,r,'Closed, each row is one location and its weekly TOTAL completions (evals + installs), shaded green above the '
        'band and red below it. Open a location for the split by week plus its own chart: the shaded bands '
        'are that location\'s Pace zones \u2014 green above Pace, amber then red below.',2+WOW_N)
r+=1

loc_region={}
for loc in allloc:
    sub=a[(a['Eval Technician'].astype(str).str.strip()==loc)|
          (a['Install Technician'].astype(str).str.strip()==loc)]
    m=sub['region'].mode()
    loc_region.setdefault(m.iat[0] if len(m) else 'Unassigned',[]).append(loc)
# Master location list, flattened -- every location that appears anywhere on
# the Locations tab. Every per-location visualization tab (Boots on Ground,
# Calls, Upload) enumerates from THIS, not from "whichever locations happen
# to have this week's data for this specific side" -- otherwise a location
# with eval work but no install work this week silently vanishes from the
# install half of a tab instead of showing a real, visible "no data" row.
ALLLOC_MASTER=sorted(set(l for locs in loc_region.values() for l in locs))
LOC_TO_REGION={l:reg for reg,locs in loc_region.items() for l in locs}

a['_estrip']=a['Eval Technician'].astype(str).str.strip()
a['_istrip']=a['Install Technician'].astype(str).str.strip()
_EV_PIVOT=a[a['edone']].groupby(['_estrip','ewk']).size().to_dict()
_INST_PIVOT=a[a['idone']].groupby(['_istrip','iwk']).size().to_dict()
_LF_PIVOT=a.groupby(['_istrip','iwk'])['lf'].sum().to_dict()
_RFQ_PIVOT=a.groupby(['_estrip','cro']).size().to_dict()
_PO_PIVOT=a.groupby(['_istrip','pow']).size().to_dict()
_EC_PIVOT=a.groupby(['_estrip','ewk'])['ec'].apply(list).to_dict()
_IC_PIVOT=a.groupby(['_istrip','iwk'])['ic'].apply(list).to_dict()

def loc_week(loc,w,what):
    """Fast lookup against precomputed pivots -- was a full-dataframe scan (and for
    rfq/po, a full-dataframe datetime re-parse) on every single call, which is
    fine called a few dozen times but becomes minutes of runtime once a 6-month
    view calls it thousands of times."""
    if what=='eval': return int(_EV_PIVOT.get((loc,w),0))
    if what=='inst': return int(_INST_PIVOT.get((loc,w),0))
    if what=='lf':   return int(_LF_PIVOT.get((loc,w),0) or 0)
    if what=='rfq':  return int(_RFQ_PIVOT.get((loc,w),0))
    if what=='po':   return int(_PO_PIVOT.get((loc,w),0))
    vals=[v for v in (_EC_PIVOT.get((loc,w),[])+_IC_PIVOT.get((loc,w),[])) if pd.notna(v)]
    return round(float(sum(vals)/len(vals)),1) if vals else None

def loc_pace(loc,what='both'):
    if what=='both':
        tot=pd.Series([loc_week(loc,w,'eval')+loc_week(loc,w,'inst') for w in BWEEKS])
    else:
        tot=pd.Series([loc_week(loc,w,what) for w in BWEEKS])
    nz=tot[tot>0]
    if len(nz)<2: return None
    tot=tot[nz.index.min():]
    mu=float(tot.mean()); sd=float(tot.std(ddof=1))
    return dict(mu=mu,sd=sd,lo2=max(0,mu-2*sd),lo1=max(0,mu-sd),
                hi1=mu+sd,hi2=mu+2*sd,n=len(tot))

for reg in REGIONS+['Unassigned']:
    if reg not in loc_region: continue
    band(ws,r,f'   {reg}',2+WOW_N); r+=1
    ranked=sorted(loc_region[reg],
                  key=lambda l:-sum(loc_week(l,w,'eval')+loc_week(l,w,'inst') for w in WEEKS))
    for loc in ranked:
        ev=[loc_week(loc,w,'eval') for w in WEEKS]
        iv=[loc_week(loc,w,'inst') for w in WEEKS]
        if not (sum(ev)+sum(iv)): continue
        pe=loc_pace(loc,'eval'); pi=loc_pace(loc,'inst')
        # ---- ONE summary row: location name only, no numbers, until expanded ----
        head_r=r
        c=ws.cell(row=r,column=1,value=loc); c.font=BD; c.fill=P(LBLU3); c.alignment=LFT
        for cc in range(2,3+WOW_N): ws.cell(row=r,column=cc).fill=P(LBLU3)
        for cc in range(1,3+WOW_N): ws.cell(row=r,column=cc).border=BOX
        ws.row_dimensions[r].outlineLevel=1; ws.row_dimensions[r].collapsed=True
        r+=1
        blk=r   # everything from here down is hidden until + is clicked
        for mi,(lbl,vals,st_,colr) in enumerate([('Evals completed',ev,pe,AA_RED),
                                                 ('Installs completed',iv,pi,AA_BLUE)]):
            c2=ws.cell(row=r,column=2,value='\u25cf  '+lbl+
                       (f'   (Pace {st_["mu"]:.0f}/wk)' if st_ else ''))
            c2.font=Font(name=FN,size=11,bold=True,color=AA_RED if mi==0 else AA_BLUE)
            c2.fill=P(T_EVAL if mi==0 else T_INST); c2.alignment=LFT
            for j,v in enumerate(vals,3):
                cel=ws.cell(row=r,column=j,value=v); cel.font=BD; cel.alignment=CTR
                cel.fill=P(T_EVAL if mi==0 else T_INST); cel.border=BOX
                if st_:
                    if   v>st_['hi2']: cel.fill=P(GRN2_Z)
                    elif v<st_['lo2']: cel.fill=P(RED_Z)
            for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
            ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True
            r+=1
        # ---- keep-up rate rows: can this location complete what came in? ----
        _rfq_wk=[loc_week(loc,w,'rfq') for w in WEEKS]
        _po_wk =[loc_week(loc,w,'po')  for w in WEEKS]
        for lbl,vals,fmt in [('RFQs received',_rfq_wk,'0'),
                             ('Eval keep-up %',[round(100*_e/_r,0) if _r else None for _r,_e in zip(_rfq_wk,ev)],'0"%"'),
                             ('POs received',_po_wk,'0'),
                             ('Install keep-up %',[round(100*_i/_p,0) if _p else None for _p,_i in zip(_po_wk,iv)],'0"%"')]:
            ws.cell(row=r,column=2,value='      '+lbl).font=RG
            mtint(ws,r,2,'RFQs received' if 'Eval' in lbl or 'RFQ' in lbl else 'POs received')
            for j,v in enumerate(vals,3):
                cel=ws.cell(row=r,column=j,value=v if v is not None else '\u2014')
                cel.font=SM; cel.alignment=CTR; cel.border=BOX
                if isinstance(v,(int,float)):
                    cel.number_format=fmt
                    if '%' in lbl: cel.fill=P(KPI_R) if v<85 else (P(KPI_A) if v<95 else P(GRN_Z))
            for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
            r+=1
        # ---- detail rows ----
        for lbl,key,fmt in [('Lineal feet installed','lf','#,##0'),
                            ('Avg call attempts','calls','0.0')]:
            ws.cell(row=r,column=2,value='      '+lbl).font=RG
            mtint(ws,r,2,lbl)
            _t2=MEASURE_TINT.get(lbl)
            if _t2:
                for _cc in range(2,3+WOW_N): ws.cell(row=r,column=_cc).fill=P(_t2)
            for j,w in enumerate(WEEKS,3):
                v=loc_week(loc,w,key)
                cel=ws.cell(row=r,column=j,value=v if v is not None else '\u2014')
                cel.font=SM; cel.alignment=CTR; cel.border=BOX
                if isinstance(v,(int,float)): cel.number_format=fmt
            for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
            r+=1
        while r-blk < BLOCK_ROWS: r+=1
        for rr in range(blk,r):
            ws.row_dimensions[rr].outlineLevel=2; ws.row_dimensions[rr].hidden=True
        # ---- chart series in hidden columns: evals + installs, each with its Pace ----
        rr=blk; srows=[]
        for nm,vals in [('Evals',ev),('Evals Pace',[round(pe['mu'],2)]*WOW_N if pe else [0]*WOW_N),
                        ('Installs',iv),('Installs Pace',[round(pi['mu'],2)]*WOW_N if pi else [0]*WOW_N)]:
            ws.cell(row=rr,column=DC,value=nm)
            for j,v in enumerate(vals,DC+1): ws.cell(row=rr,column=j,value=v)
            srows.append(rr); rr+=1
        lc=LineChart()
        lc.add_data(Reference(ws,min_col=DC+1,max_col=DC+WOW_N,min_row=srows[0],max_row=srows[-1]),
                    from_rows=True)
        lc.set_categories(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=4))
        for k,nm in enumerate(['Evals completed','Evals Pace','Installs completed','Installs Pace']):
            lc.series[k].tx=SeriesLabel(v=nm)
        style_line(lc,[(CHART_TINT['eval'],None,'diamond',28000),(AA_RED,'lgDash',None,25400),
                       (CHART_TINT['install'],None,'diamond',28000),(AA_BLUE,'lgDash',None,25400)],
                   loc,'')                      # short title: the row labels give the context
        lc.height=9.5; lc.width=17
        # no legend: the two rows to the left are already tinted red for evals and
        # blue for installs, so a key here would only crowd the date labels
        lc.legend=None
        lc.y_axis.title=None
        lc.x_axis.title=loc                 # location name sits under the chart
        vertical_labels(lc.x_axis,700)      # dates read straight down, never stacked
        # anchor over the HIDDEN rows only, so the chart vanishes entirely when closed
        lc.anchor=TwoCellAnchor(editAs='twoCell',
            _from=AnchorMarker(col=2+WOW_N,colOff=0,row=blk,rowOff=0),
            to=AnchorMarker(col=2+WOW_N+9,colOff=0,row=r-2,rowOff=0))
        ws.add_chart(lc)
        # ---- two small flow pies beside the line chart: intake vs completion, THIS WEEK ----
        _this_rfq=_rfq_wk[-1]; _this_eval=ev[-1]; _this_po=_po_wk[-1]; _this_inst=iv[-1]
        rr+=1
        ws.cell(row=rr,column=DC,value='RFQs received'); ws.cell(row=rr,column=DC+1,value=_this_rfq)
        ws.cell(row=rr+1,column=DC,value='Evals completed'); ws.cell(row=rr+1,column=DC+1,value=_this_eval)
        ws.cell(row=rr+2,column=DC,value='POs received'); ws.cell(row=rr+2,column=DC+1,value=_this_po)
        ws.cell(row=rr+3,column=DC,value='Installs completed'); ws.cell(row=rr+3,column=DC+1,value=_this_inst)
        _pie_col = 2+WOW_N+11          # clear of the line chart, which ends at col 2+WOW_N+9
        if _this_rfq or _this_eval:
            _p1=PieChart(); _p1.height=9.5; _p1.width=12
            _p1.title='Eval flow this week'
            _p1.add_data(Reference(ws,min_col=DC+1,min_row=rr,max_row=rr+1))
            _p1.set_categories(Reference(ws,min_col=DC,min_row=rr,max_row=rr+1))
            _p1.series[0].data_points=[_DP2(idx=0,spPr=GraphicalProperties(solidFill=CHART_TINT['eval'])),
                                       _DP2(idx=1,spPr=GraphicalProperties(solidFill=AA_RED))]
            _p1.series[0].dLbls=DataLabelList(showVal=True,dLblPos='ctr',
                showCatName=False,showSerName=False,showPercent=False,showLegendKey=False,showBubbleSize=False)
            _p1.series[0].dLbls.txPr=RichText(bodyPr=RichTextProperties(),
                p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1000,b=True,solidFill='FFFFFF')),
                             endParaRPr=CharacterProperties(sz=1000,b=True,solidFill='FFFFFF'))])
            _p1.legend.position='b'
            _p1.anchor=TwoCellAnchor(editAs='twoCell',
                _from=AnchorMarker(col=_pie_col,colOff=0,row=blk,rowOff=0),
                to=AnchorMarker(col=_pie_col+6,colOff=0,row=blk+9,rowOff=0))
            ws.add_chart(_p1)
        if _this_po or _this_inst:
            _p2=PieChart(); _p2.height=9.5; _p2.width=12
            _p2.title='Install flow this week'
            _p2.add_data(Reference(ws,min_col=DC+1,min_row=rr+2,max_row=rr+3))
            _p2.set_categories(Reference(ws,min_col=DC,min_row=rr+2,max_row=rr+3))
            _p2.series[0].data_points=[_DP2(idx=0,spPr=GraphicalProperties(solidFill=CHART_TINT['install'])),
                                       _DP2(idx=1,spPr=GraphicalProperties(solidFill=AA_BLUE))]
            _p2.series[0].dLbls=DataLabelList(showVal=True,dLblPos='ctr',
                showCatName=False,showSerName=False,showPercent=False,showLegendKey=False,showBubbleSize=False)
            _p2.series[0].dLbls.txPr=RichText(bodyPr=RichTextProperties(),
                p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1000,b=True,solidFill='FFFFFF')),
                             endParaRPr=CharacterProperties(sz=1000,b=True,solidFill='FFFFFF'))])
            _p2.legend.position='b'
            _p2.anchor=TwoCellAnchor(editAs='twoCell',
                _from=AnchorMarker(col=_pie_col+7,colOff=0,row=blk,rowOff=0),
                to=AnchorMarker(col=_pie_col+13,colOff=0,row=blk+9,rowOff=0))
            ws.add_chart(_p2)
    r+=1

anchor=r+2
for metric,colr in [('Eval',AA_RED),('Install',AA_BLUE)]:
    rows=INTROWS[metric][:2]                     # IN and OUT only
    lc=LineChart()
    lc.add_data(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=rows[0],max_row=rows[1]),from_rows=True)
    lc.set_categories(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=4))
    lc.series[0].tx=SeriesLabel(strRef=StrRef(f"'Trends'!$A${rows[0]}"))
    lc.series[1].tx=SeriesLabel(strRef=StrRef(f"'Trends'!$A${rows[1]}"))
    style_line(lc,[('7F7F7F',None,'square',22000),(colr,None,'diamond',28000)],
               f'{metric} intake vs completed','')
    ws.add_chart(lc,f'A{anchor}'); anchor+=17

for met,colr in [('Evals',AA_RED),('Installs',AA_BLUE)]:
    rows=NATROWS[met]
    lc=LineChart()
    lc.add_data(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=rows[0],max_row=rows[-1]),from_rows=True)
    lc.set_categories(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=4))
    for k,rw in enumerate(rows):
        lc.series[k].tx=SeriesLabel(strRef=StrRef(f"'Trends'!$A${rw}"))
    style_line(lc,[(colr,None,'diamond',28000),
                   ('1A1A1A','lgDash',None,31750),
                   (AA_SILVER,'sysDot',None,12700),
                   (AA_SILVER,'sysDot',None,12700)],
               f'{met} vs Pace','')
    ws.add_chart(lc,f'A{anchor}'); anchor+=17

for met in ['Evals ✓','Installs ✓','Avg Calls']:      # Eval quote-sent (d) graph removed on request
    rows=rowmap[met]
    lc=LineChart()
    # region rows are no longer contiguous, so add each one separately
    for rw in rows:
        lc.add_data(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=rw,max_row=rw),from_rows=True)
    lc.set_categories(Reference(ws,min_col=3,max_col=2+WOW_N,min_row=4))
    for k,reg in enumerate(REGIONS):
        lc.series[k].tx=SeriesLabel(v=reg)
    style_line(lc,[(AA_BLUE,None,'diamond',28000),
                   (AA_RED,None,'circle',22000),
                   (AA_SILVER,None,'circle',22000)],
               f'{met} by region','')
    CAP={'Evals ✓':'Evaluations completed each week. Divergence between the three lines is normal — '
                   'East and West carry far more volume than Central. Watch the SHAPE, not the gap.',
         'Installs ✓':'Installs completed each week. A dip in a holiday week is expected; a dip that '
                      'persists two or more weeks is worth a call.',
         'Eval quote-sent (d)':'Days from evaluation to quote sent. LOWER IS BETTER. Rising lines mean '
                          'paperwork is slowing down even if completion counts hold steady.',
         'Avg Calls':'Average call attempts to reach the veteran. Rising lines mean scheduling is '
                     'getting harder — often the leading indicator of a slowdown next week.'}
    caption(ws,anchor-1,CAP[met],2+WOW_N)
    ws.add_chart(lc,f'A{anchor}'); anchor+=17

# ==================================================== ALL LOCATIONS BY REGION
EVAL_LAG_EXCLUSIONS=[]   # (location, project_id, lag_days, is_confirmed_reeval)
ws=wb.create_sheet('Locations')
titles(ws,'Locations — this week measured against each location\'s own history','Region bands, region subtotals in bold. Left half is this week; right half is the historical comparison.')
COLS=['Location','Channel','Evals ✓','Evals pend','RFQs received',
      'Installs ✓','Installs pend','POs received','Install Pace/wk','This week vs Pace',
      'Total LF','Avg Calls','Avg Eval Upload (d)','Install upload (d)',
      'Eval time on site','Install time on site','Avg Eval BOG (d)','Avg Install BOG (d)',
      'Re-evals / Active','Zone this week','Tier','Consistency','Band (hidden)']
key_block(ws,4,len(COLS)+2,ZONE_KEY+[(IC_SC,'Service Center',CHAN_GREEN,None),
                                     (IC_DEALER,'Dealer',CHAN_AMBER,None),
                                     (IC_NONE,'not measured',LRED3,None)],
          'KEY \u2014 zone / channel')
hdr(ws,4,COLS,[38,15,10,11,14,11,12,14,24,15,11,11,13,14,16,17,15,16,14,14,7,12,14])
for _c in ('T','U','V','W'): ws.column_dimensions[_c].hidden=True   # zone text / tier / consistency / band
caption(ws,3,"Left half is THIS WEEK. Right half compares it to that location's own 48-week history: "
        "Pace is its normal weekly output, Band is Pace ±2 sigma, Zone says where this week landed — green "
        "above Pace is good, amber and red below. Consistency near 1.0 is ordinary week-to-week noise; above "
        "1.5 the location is genuinely erratic. Time on site reads 'insufficient' where too few crews logged it.",
        len(COLS))
ws.freeze_panes='A5'
r=5
for reg in REGIONS+['Unassigned']:
    locs=set(loc_region.get(reg,[]))   # dominant/parent region always -- an
    # outlier project (e.g. one Biloxi eval on an otherwise-East location) stays
    # in that location's own numbers, it just doesn't relist the location itself
    # under a different region band.
    if not locs: continue
    band(ws,r,f'◆  {reg}',len(COLS)); r+=1
    _region_start=r
    rows=[]
    for loc in locs:
        eo,ep_,e=stat(EW,'edone','Eval Technician',loc); io,ip_,i=stat(IW,'idone','Install Technician',loc)
        rows.append((io,loc,eo,ep_,io,ip_,e,i))
    tot=[0]*6; totlf=0
    for _,loc,eo,ep_,io,ip_,e,i in sorted(rows,reverse=True,key=lambda t:(t[0],t[1])):
        ul=e['ulag'].dropna()
        # Active book for this location -- not just this week's worked evals/
        # installs, but every currently active project assigned to it, same
        # population the new Re-Evaluations tab uses.
        _loc_active=a[((a['_estrip']==loc)|(a['_istrip']==loc))&a['_is_open']]
        _active_n=len(_loc_active)
        _reeval_n=int(_loc_active['_is_reeval'].sum())
        _il=(i[i['_pid'].isin(upl.index)]['ulag_install_real'].dropna()
             if (USE_REAL_UPLOAD and upl is not None) else pd.Series(dtype=float))
        _el_sub=(e[e['_pid'].isin(eup.index)] if eup is not None else e.iloc[0:0])
        # A real upload dated well before the eval date almost always means a
        # re-eval overwrote the date field while the upload still belongs to
        # the ORIGINAL visit -- not a real measurement of anything. -1 day
        # allows for normal rounding/timezone slack; anything past that gets
        # pulled out of the average and named explicitly rather than either
        # silently dragging the mean or silently vanishing.
        _el=_el_sub.loc[_el_sub['ulag_eval_real']>=-1,'ulag_eval_real'].dropna()
        _el_out_rows=_el_sub.loc[_el_sub['ulag_eval_real']<-1]
        for _,_orow in _el_out_rows.iterrows():
            _note_out=str(_orow.get('Most Recent Timeline Note') or '')
            _is_reeval='RE-EVAL' in _note_out.upper() or 'REEVAL' in _note_out.upper()
            EVAL_LAG_EXCLUSIONS.append((loc,_orow['Project ID'],round(float(_orow['ulag_eval_real']),1),_is_reeval))
        st=pace_stats(loc,'Install')
        zname,zfill=zone_of(io,st)
        _cro=pd.to_datetime(a['Created On'],errors='coerce').apply(wk_of)
        _pow=pd.to_datetime(a['PO Received Date'],errors='coerce').apply(wk_of)
        _rq=int(((_cro==WEEK)&(a['Eval Technician'].astype(str).str.strip()==loc)).sum())
        _po=int(((_pow==WEEK)&(a['Install Technician'].astype(str).str.strip()==loc)).sum())
        erep=(_rq/eo) if eo else None
        irep=(_po/io) if io else None
        vals=[loc,'\u25a0 Service Center' if is_sc(loc) else '\u25a1 Dealer',
              eo, ep_, _rq,
              io, ip_, _po,
              round(st['mean'],1) if st else None,
              # Symbol in place of "vs" -- same info, more compact. The full
              # wording (e.g. "On pace", "Caution") stays available in the
              # hidden Zone this week column for anyone who wants it spelled out.
              (f"{io} {zname.split(' ',1)[0] if zname else ''} {st['mean']:.1f}"
               if st and st['tier']!='C' else None),
              int(i['lf'].sum()),
              round(float(pd.concat([e['ec'],i['ic']]).dropna().mean()),1) if pd.concat([e['ec'],i['ic']]).notna().any() else None,
              (round(float(_el.mean()),1) if len(_el) else ('n/a' if eup is None else None)),
              round(float(_il.mean()),1) if len(_il) else ('n/a' if not (USE_REAL_UPLOAD and upl is not None) else None),
              tos_for(loc,'Eval'), tos_for(loc,'Install'),
              (round(float(pd.to_numeric(e['ebog_adj'],errors='coerce').dropna().mean()),1)
               if pd.to_numeric(e['ebog_adj'],errors='coerce').notna().any() else None),
              (round(float(pd.to_numeric(i['ibog_adj'],errors='coerce').dropna().mean()),1)
               if pd.to_numeric(i['ibog_adj'],errors='coerce').notna().any() else None),
              # Separator is an en-dash, NOT a slash: Excel auto-parses "1/140"
              # as a date and slaps a green-triangle warning on every cell. The
              # en-dash keeps it unambiguously text and stops the false flags.
              (f"{int(_reeval_n)} \u2013 {int(_active_n)}" if _active_n else None),
              zname, st['tier'] if st else None,
              (round(st['cons'],2) if st and st['cons'] else None),
              (f"{st['lo2']:.0f}\u2013{st['hi2']:.0f}" if st and st['tier']!='C' else None)]
        for cc,v in enumerate(vals,1):
            c=ws.cell(row=r,column=cc,value=v if v is not None else '—')
            _t=MEASURE_TINT.get(COLS[cc-1]) if cc<=len(COLS) else None
            if _t: c.fill=P(_t)
            c.font=BD if cc==1 else RG; c.border=BOX
            c.alignment=LFT if cc<=2 else CTR
            if cc==9 and v is not None: c.number_format='0.0'
            if cc in (13,14) and isinstance(v,float): c.number_format='0.0'
            # Re-evals / Active is a ratio string, not a number or a date --
            # force explicit text format so Excel never tries to reinterpret it.
            if cc==19: c.number_format='@'
        ws.cell(row=r,column=2).fill=P(CHAN_GREEN) if is_sc(loc) else P(CHAN_AMBER)
        for cc in (4,7):
            if ws.cell(row=r,column=cc).value: ws.cell(row=r,column=cc).fill=P(KPI_A)
        for _rc in (5,8):
            _cel=ws.cell(row=r,column=_rc); _cel.alignment=CTR
        if zfill:
            ws.cell(row=r,column=9).fill=P(zfill); ws.cell(row=r,column=9).font=BD
            ws.cell(row=r,column=16).fill=P(zfill)
        if st and st['cons'] and st['cons']>1.5: ws.cell(row=r,column=14).fill=P(KPI_A)
        for _tc in (14,15):
            if ws.cell(row=r,column=_tc).value in ('insufficient','n/a'):
                ws.cell(row=r,column=_tc).font=NOTE; ws.cell(row=r,column=_tc).fill=P(LRED3)
        # no row banding: the column tints already separate the measures
        for _cc,_h in enumerate(COLS,1): mtint(ws,r,_cc,_h)
        # Six real values, one per data column (Evals pend/RFQs/Installs pend/POs
        # were being silently skipped before -- the subtotal row's RFQs-received
        # and Installs-completed columns were showing each other's numbers).
        tot=[tot[0]+eo,tot[1]+ep_,tot[2]+_rq,tot[3]+io,tot[4]+ip_,tot[5]+_po]; totlf+=int(i['lf'].sum())
        r+=1
    # ---- region subtotal: banded, ruled top and bottom, clearly not a location ----
    _med=Side(style='medium',color=AA_BLUE)
    for cc in range(1,len(COLS)+1):
        cel=ws.cell(row=r,column=cc)
        cel.fill=P(LBLU1)
        cel.border=Border(left=thin,right=thin,top=_med,bottom=_med)
    for cc,v in ((1,f'{reg.upper()}  \u2014  SUBTOTAL'),(3,tot[0]),(4,tot[1]),
                 (5,tot[2]),(6,tot[3]),(7,tot[4]),(8,tot[5]),(11,totlf)):
        c=ws.cell(row=r,column=cc,value=v)
        c.font=Font(name=FN,size=11,bold=True,color=AA_BLUE)
        c.alignment=LFT if cc==1 else CTR
        if cc==11: c.number_format='#,##0'
    ws.cell(row=r,column=2,value=f'{len(rows)} locations').font=Font(name=FN,size=10,italic=True,color=AA_BLUE)
    ws.cell(row=r,column=2).alignment=CTR
    ws.row_dimensions[r].height=20
    _region_end=r-1        # last real location row, before this subtotal
    if _region_end>=_region_start:
        # Share of region total, not relative min/max ranking within the
        # region -- a location's bar should read as "this many of the
        # region's total", matching the same convention already used on
        # Trends (share-of-week bars). end_value is the region subtotal
        # itself (tot[0] for Evals, tot[3] for Installs), not the highest
        # individual location's count.
        if tot[0]:
            ws.conditional_formatting.add(f'C{_region_start}:C{_region_end}',
                DataBarRule(start_type='num',start_value=0,end_type='num',end_value=tot[0],
                            color=AA_RED,showValue=True))
        if tot[3]:
            ws.conditional_formatting.add(f'F{_region_start}:F{_region_end}',
                DataBarRule(start_type='num',start_value=0,end_type='num',end_value=tot[3],
                            color=AA_BLUE,showValue=True))
    r+=1
    ws.row_dimensions[r].height=8          # breathing room before the next region
    r+=1

if EVAL_LAG_EXCLUSIONS:
    band(ws,r,f'\u26a0  EXCLUDED FROM AVG EVAL UPLOAD \u2014 {len(EVAL_LAG_EXCLUSIONS)} project(s), likely re-eval date mismatch',len(COLS)); r+=1
    hdr(ws,r,['Location','Project','Apparent lag (d)','Confirmed by note text?','','','','',''])
    r+=1
    for _loc,_pid,_lag,_confirmed in EVAL_LAG_EXCLUSIONS:
        ws.cell(row=r,column=1,value=_loc).font=RG
        ws.cell(row=r,column=2,value=_pid).font=BD
        ws.cell(row=r,column=3,value=_lag).alignment=CTR
        ws.cell(row=r,column=4,value='Yes -- note says RE-EVAL' if _confirmed else 'No -- flagged on threshold only').alignment=LFT
        for cc in range(1,10): ws.cell(row=r,column=cc).border=BOX
        r+=1
    caption(ws,r,'A real upload timestamp dated more than 1 day before the eval date almost always means a '
            're-eval overwrote the date field while the upload still belongs to the original visit -- not a real '
            'measurement of anything. These are pulled out of every location\'s Avg Eval Upload average above; '
            'the raw number is still visible on Completions Detail for anyone who wants the specific value.',len(COLS))
    r+=2

# ==================================================== BY LOCATION (WEEKLY)
ws=wb.create_sheet('By Location (Weekly)')
titles(ws,'By Location — Weekly','Collapsible per-location weekly grid with a mini trend chart each.')
c=ws.cell(row=2,column=6,value='Evals ✓'); c.fill=P(AA_RED); c.font=HDW; c.alignment=CTR
c=ws.cell(row=2,column=7,value='Installs ✓'); c.fill=P(AA_BLUE); c.font=HDW; c.alignment=CTR
hdr(ws,4,['Location / Metric']+[w.strftime('%-m/%-d') for w in WEEKS],[38]+[9]*WOW_N)
ws.freeze_panes='A5'
ws.sheet_properties.outlinePr.summaryBelow=False
r=5
allloc=set(a[a['ewk'].isin(WEEKS)]['Eval Technician'].dropna().astype(str).str.strip()) | \
        set(a[a['iwk'].isin(WEEKS)]['Install Technician'].dropna().astype(str).str.strip())
byreg={}
for loc in allloc:
    sub=a[(a['Eval Technician'].astype(str).str.strip()==loc)|(a['Install Technician'].astype(str).str.strip()==loc)]
    byreg.setdefault(sub['region'].mode().iat[0] if len(sub['region'].mode()) else 'Unassigned',[]).append(loc)
for reg in REGIONS+['Unassigned']:
    if reg not in byreg: continue
    band(ws,r,f'◆  {reg}',1+WOW_N); r+=1
    for loc in sorted(byreg[reg]):
        c=ws.cell(row=r,column=1,value=loc); c.font=BD; c.fill=P(LBLU3)
        for cc in range(2,2+WOW_N): ws.cell(row=r,column=cc).fill=P(LBLU3)
        ws.row_dimensions[r].outlineLevel=1; ws.row_dimensions[r].collapsed=True
        head=r; r+=1
        base=r
        for met in ['Evals ✓','Installs ✓']:
            ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True
            c=ws.cell(row=r,column=1,value='   '+met); c.font=RG
            for j,w in enumerate(WEEKS,2):
                if met=='Evals ✓':
                    v=int(a[(a['ewk']==w)&(a['Eval Technician'].astype(str).str.strip()==loc)]['edone'].sum())
                else:
                    v=int(a[(a['iwk']==w)&(a['Install Technician'].astype(str).str.strip()==loc)]['idone'].sum())
                cc=ws.cell(row=r,column=j,value=v); cc.alignment=CTR; cc.border=BOX
            r+=1
        lc=LineChart(); lc.style=2; lc.height=5.5; lc.width=14; lc.title=f'{loc} — weekly'
        lc.add_data(Reference(ws,min_col=2,max_col=1+WOW_N,min_row=base,max_row=base+1),from_rows=True)
        lc.set_categories(Reference(ws,min_col=2,max_col=1+WOW_N,min_row=4))
        for k,srs in enumerate(lc.series):
            srs.graphicalProperties=GraphicalProperties()
            srs.graphicalProperties.line=LineProperties(solidFill=AA_RED if k==0 else AA_BLUE,w=28000)
            srs.marker=Marker(symbol='circle',size=4); srs.smooth=False
            srs.tx=SeriesLabel(v=("Evals ✓" if k==0 else "Installs ✓"))
        lc.legend.position='b'
        ws.add_chart(lc,f'{L(3+WOW_N)}{head}')
        r+=4

# ==================================================== TRENDS EXPANDED (6-month, weekly)
WOW26=26
WEEKS26=[WEEK-pd.Timedelta(weeks=i) for i in range(WOW26-1,-1,-1)]

ws=wb.create_sheet('6 Month Trends')
ws.sheet_view.showGridLines=False
titles(ws,'6 Month Trends \u2014 the same locations, six months of weeks',
       f'{WOW26} weeks ending {WEEK.strftime("%b %-d")}. Same measures and layout as Trends, just a '
       'longer window so a slower-moving pattern is visible in one place.')
ws.freeze_panes='C5'
XMETS=['Evals \u2713','Installs \u2713','Total LF','Avg Calls']
r=5
xrowmap={}
XREGION_TOP=r
for reg in REGIONS:
    band(ws,r,f'   {reg}',2+WOW26); r+=1
    for met in XMETS:
        ws.cell(row=r,column=1,value='').font=RG
        ws.cell(row=r,column=2,value=met).font=BD
        mtint(ws,r,2,met)
        _t=MEASURE_TINT.get(met)
        if _t:
            for _cc in range(2,3+WOW26): ws.cell(row=r,column=_cc).fill=P(_t)
        for j,w in enumerate(WEEKS26,3):
            e=a[(a['ewk']==w)&(a['region']==reg)]; i=a[(a['iwk']==w)&(a['region']==reg)]
            if   met=='Evals \u2713':    v=int(e['edone'].sum())
            elif met=='Installs \u2713': v=int(i['idone'].sum())
            elif met=='Total LF':    v=int(i['lf'].sum())
            else:                    v=round(float(pd.concat([e['ec'],i['ic']]).dropna().mean()),1) if len(e)+len(i) else 0
            c=ws.cell(row=r,column=j,value=v); c.font=RG; c.alignment=CTR; c.border=BOX
        for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
        xrowmap.setdefault(met,[]).append(r); r+=1
    r+=1
hdr(ws,4,['Region / Location','Measure']+[w.strftime('%-m/%-d') for w in WEEKS26],[38,17]+[9]*WOW26)

# share-of-week data bars, same logic as the 5-week Trends tab
for met in ('Evals \u2713','Installs \u2713'):
    _rrows=xrowmap[met]
    for j in range(3,3+WOW26):
        _vals=[ws.cell(row=rr,column=j).value or 0 for rr in _rrows]
        _total=sum(_vals)
        if _total<=0: continue
        _col=L(j); _sqref=' '.join(f'{_col}{rr}' for rr in _rrows)
        ws.conditional_formatting.add(_sqref, DataBarRule(
            start_type='num', start_value=0, end_type='num', end_value=_total,
            color=AA_RED if met=='Evals \u2713' else AA_BLUE, showValue=True))

# one chart per measure, all three regions, across all 26 weeks -- 2x2 grid,
# same reasoning as the 5-week Trends tab: keeps the block compact instead of
# stacking 4 charts tall and reaching down into the table content below.
_xgrid=[(0,0),(14,0),(0,16),(14,16)]
for _xgi,met in enumerate(XMETS):
    _rows=xrowmap[met]
    _mc=LineChart()
    for _rw in _rows:
        _mc.add_data(Reference(ws,min_col=3,max_col=2+WOW26,min_row=_rw,max_row=_rw),from_rows=True)
    _mc.set_categories(Reference(ws,min_col=3,max_col=2+WOW26,min_row=4))
    for _k,_reg in enumerate(REGIONS): _mc.series[_k].tx=SeriesLabel(v=_reg)
    style_line(_mc,REG_SPEC,f'{met} \u2014 by region, 6 months','')
    vertical_labels(_mc.x_axis,700)
    _mc.height=9; _mc.width=22
    _xcoff,_xroff=_xgrid[_xgi]
    _mc.anchor=TwoCellAnchor(editAs='oneCell',
        _from=AnchorMarker(col=2+WOW26+1+_xcoff,colOff=0,row=XREGION_TOP-1+_xroff,rowOff=0),
        to=AnchorMarker(col=2+WOW26+13+_xcoff,colOff=0,row=XREGION_TOP+13+_xroff,rowOff=0))
    ws.add_chart(_mc)
r+=1

# ---- nationwide intake vs completed, by week, 6 months ----
band(ws,r,'\u25c6  Intake \u2014 total RFQs and POs received vs completed, 6 months',2+WOW26); r+=1
caption(ws,r,'Same keep-up logic as the weekly Trends tab: completed divided by received. Below 100% '
        'that week fell behind; above 100% it caught up.',2+WOW26); r+=1
for metric,incol,inlbl,outlbl,done,wkcol in [
        ('Eval','cro','RFQs received (IN)','Evals completed (OUT)','edone','ewk'),
        ('Install','pow','POs received (IN)','Installs completed (OUT)','idone','iwk')]:
    ins=[int((a[incol]==w).sum()) for w in WEEKS26]
    outs=[int(a[a[wkcol]==w][done].sum()) for w in WEEKS26]
    for lbl,vals,fmt in [(inlbl,ins,'0'),(outlbl,outs,'0'),
                         ('Keep-up rate',[(o/i) if i else None for i,o in zip(ins,outs)],'0%')]:
        ws.cell(row=r,column=1,value=lbl).font=BD if 'Keep-up' not in lbl else RG
        ws.cell(row=r,column=2,value=metric).font=RG
        for j,v in enumerate(vals,3):
            cel=ws.cell(row=r,column=j,value=v if v is not None else '\u2014')
            cel.font=RG; cel.alignment=CTR; cel.border=BOX; cel.number_format=fmt
            if 'Keep-up' in lbl and isinstance(v,float):
                cel.fill=P(KPI_R) if v<0.85 else (P(KPI_A) if v<0.95 else P(GRN_Z))
        for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
        r+=1
    r+=1

# ---- per-location, by week, 6 months, collapsible ----
band(ws,r,'\u25c6  By Location \u2014 click the + to expand, 6 months of weeks',2+WOW26); r+=1
caption(ws,r,'Closed by default, only the location name shows until expanded -- same as the weekly '
        'Trends tab. Pies show the most recent week only.',2+WOW26); r+=1
ws.sheet_properties.outlinePr.summaryBelow=False
XBLOCK_ROWS=14
XDC=2+WOW26+16
for reg in REGIONS+['Unassigned']:
    if reg not in loc_region: continue
    band(ws,r,f'   {reg}',2+WOW26); r+=1
    for loc in sorted(loc_region[reg]):
        evx=[loc_week(loc,w,'eval') for w in WEEKS26]
        ivx=[loc_week(loc,w,'inst') for w in WEEKS26]
        if not (sum(evx)+sum(ivx)): continue
        pe=loc_pace(loc,'eval'); pi=loc_pace(loc,'inst')
        xhead=r
        c=ws.cell(row=r,column=1,value=loc); c.font=BD; c.fill=P(LBLU3); c.alignment=LFT
        for cc in range(2,3+WOW26): ws.cell(row=r,column=cc).fill=P(LBLU3)
        for cc in range(1,3+WOW26): ws.cell(row=r,column=cc).border=BOX
        ws.row_dimensions[r].outlineLevel=1; ws.row_dimensions[r].collapsed=True
        r+=1
        xblk=r
        for mi,(lbl,vals,colr,pace) in enumerate([('Evals completed',evx,AA_RED,pe),
                                                   ('Installs completed',ivx,AA_BLUE,pi)]):
            c2=ws.cell(row=r,column=2,value='\u25cf  '+lbl); c2.font=Font(name=FN,size=11,bold=True,color=colr)
            c2.fill=P(T_EVAL if mi==0 else T_INST); c2.alignment=LFT
            for j,v in enumerate(vals,3):
                cel=ws.cell(row=r,column=j,value=v); cel.font=BD; cel.alignment=CTR
                cel.fill=P(T_EVAL if mi==0 else T_INST); cel.border=BOX
            for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
            ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True; r+=1
            # Mean (Pace) row: same number every week, right under what it's the
            # mean OF, so it's visible in the table itself and not only as the
            # invisible line the dashed chart series reads from off to the side.
            _mval=round(pace['mu'],1) if pace else None
            c3=ws.cell(row=r,column=2,value=f'      Mean (Pace, full baseline): {_mval if _mval is not None else "\u2014"}/wk')
            c3.font=Font(name=FN,size=10,italic=True,color=colr)
            c3.fill=P(T_EVAL if mi==0 else T_INST)
            for j in range(3,3+WOW26):
                cel=ws.cell(row=r,column=j,value=_mval)
                cel.font=NOTE; cel.alignment=CTR; cel.fill=P(T_EVAL if mi==0 else T_INST); cel.border=BOX
                if _mval is not None: cel.number_format='0.0'
            for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
            ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True; r+=1
        rfq_x=[loc_week(loc,w,'rfq') for w in WEEKS26]
        po_x =[loc_week(loc,w,'po')  for w in WEEKS26]
        for lbl,vals,fmt in [('RFQs received',rfq_x,'0'),
                             ('Eval keep-up %',[round(100*_e/_r,0) if _r else None for _r,_e in zip(rfq_x,evx)],'0"%"'),
                             ('POs received',po_x,'0'),
                             ('Install keep-up %',[round(100*_i/_p,0) if _p else None for _p,_i in zip(po_x,ivx)],'0"%"')]:
            ws.cell(row=r,column=2,value='      '+lbl).font=RG
            mtint(ws,r,2,'RFQs received' if 'RFQ' in lbl or 'Eval' in lbl else 'POs received')
            for j,v in enumerate(vals,3):
                cel=ws.cell(row=r,column=j,value=v if v is not None else '\u2014')
                cel.font=SM; cel.alignment=CTR; cel.border=BOX
                if isinstance(v,(int,float)):
                    cel.number_format=fmt
                    if '%' in lbl: cel.fill=P(KPI_R) if v<85 else (P(KPI_A) if v<95 else P(GRN_Z))
            for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
            ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True; r+=1
        for lbl,key,fmt in [('Lineal feet installed','lf','#,##0'),('Avg call attempts','calls','0.0')]:
            vals=[loc_week(loc,w,key) for w in WEEKS26]
            ws.cell(row=r,column=2,value='      '+lbl).font=RG
            mtint(ws,r,2,lbl)
            for j,v in enumerate(vals,3):
                cel=ws.cell(row=r,column=j,value=v if v is not None else '\u2014')
                cel.font=SM; cel.alignment=CTR; cel.border=BOX
                if isinstance(v,(int,float)): cel.number_format=fmt
            for cc in (1,2): ws.cell(row=r,column=cc).border=BOX
            ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True; r+=1
        while r-xblk < XBLOCK_ROWS: r+=1
        for _rr in range(xblk,r):
            ws.row_dimensions[_rr].outlineLevel=2; ws.row_dimensions[_rr].hidden=True
        # trend line, all 26 weeks, WITH the mean/Pace dashed line -- same convention
        # as the 5-week Trends tab. Actual line is now full brand color (not the
        # pale tint) and heavier weight, since a thin pale line reads poorly across
        # 26 noisier points; the dashed Pace line is unchanged from the normal view.
        _xc2=xblk; _srows=[]
        for _nm,_vals in [('Evals',evx),('Evals Pace',[round(pe['mu'],2)]*WOW26 if pe else [0]*WOW26),
                          ('Installs',ivx),('Installs Pace',[round(pi['mu'],2)]*WOW26 if pi else [0]*WOW26)]:
            ws.cell(row=_xc2,column=XDC,value=_nm)
            for j,v in enumerate(_vals,XDC+1): ws.cell(row=_xc2,column=j,value=v)
            _srows.append(_xc2); _xc2+=1
        _xlc=LineChart()
        _xlc.add_data(Reference(ws,min_col=XDC+1,max_col=XDC+WOW26,min_row=_srows[0],max_row=_srows[-1]),from_rows=True)
        _xlc.set_categories(Reference(ws,min_col=3,max_col=2+WOW26,min_row=4))
        for _k,_nm in enumerate(['Evals completed','Evals Pace','Installs completed','Installs Pace']):
            _xlc.series[_k].tx=SeriesLabel(v=_nm)
        style_line(_xlc,[(AA_RED,None,'diamond',34000),(AA_RED,'lgDash',None,25400),
                         (AA_BLUE,None,'diamond',34000),(AA_BLUE,'lgDash',None,25400)],loc,'')
        _xlc.legend=None; _xlc.y_axis.title=None; _xlc.x_axis.title=loc
        _xlc.height=9.5; _xlc.width=22
        vertical_labels(_xlc.x_axis,700)
        _xlc.anchor=TwoCellAnchor(editAs='twoCell',
            _from=AnchorMarker(col=2+WOW26,colOff=0,row=xblk,rowOff=0),
            to=AnchorMarker(col=2+WOW26+11,colOff=0,row=r-2,rowOff=0))
        ws.add_chart(_xlc)
    r+=1

# ==================================================== COMPLETIONS DETAIL
# ==================================================== COMPLETIONS DETAIL
ws=wb.create_sheet('Completions Detail')
titles(ws,'Completions Detail — grouped by region and location','Use the +/- buttons at the left to expand. Not-uploaded rows sort to the top of each section.')
DC=['Type','Project','Customer (VAMC)','Date','Uploaded?','Days to Current','Likely Stage',
    'Eval Calls','Inst Calls','Upload Lag (d)','LF','Re-eval?','Most Recent Timeline Note']
hdr(ws,4,DC,[10,14,38,11,11,13,26,11,11,14,9,10,60])
caption(ws,3,'Days to Current: for a not-yet-uploaded row, days since the last unblocking event for '
        'that side -- RFQ received (Created On proxy) for Evals, PO Received Date for Installs. Same '
        'two-anchor convention as the analyst team\'s own Aging Report. Blank for completed rows -- '
        'nothing is aging once it\'s done. Likely Stage is a first read from data already on hand: it '
        'will sharpen once Contact Attempt Days (days actually spent calling, not call count) replaces '
        'the interim call-count check used here.',len(DC))
ws.freeze_panes='A5'
ws.sheet_properties.outlinePr.summaryBelow=False
r=5
for reg in REGIONS+['Unassigned']:
    locs=set(loc_region.get(reg,[]))   # dominant/parent region always -- an
    # outlier project (e.g. one Biloxi eval on an otherwise-East location) stays
    # in that location's own numbers, it just doesn't relist the location itself
    # under a different region band.
    if not locs: continue
    band(ws,r,f'◆  {reg}',len(DC)); r+=1
    for loc in sorted(locs):
        c=ws.cell(row=r,column=1,value=f'   {loc}'); c.font=BD
        for cc in range(1,len(DC)+1): ws.cell(row=r,column=cc).fill=P(LBLU3); ws.cell(row=r,column=cc).border=BOX
        ws.row_dimensions[r].outlineLevel=1; ws.row_dimensions[r].collapsed=True
        r+=1
        for met,W,tcol,dcol,done in [('Evals',EW,'Eval Technician','edt','edone'),
                                     ('Installs',IW,'Install Technician','idt','idone')]:
            sub=W[W[tcol].astype(str).str.strip()==loc]
            if not len(sub): continue
            nno=int((~sub[done]).sum())
            for cc in range(1,len(DC)+1): ws.cell(row=r,column=cc).fill=FSUB
            c=ws.cell(row=r,column=1,value=f'      {met}  —  {len(sub)} total, {nno} NOT uploaded'); c.font=HDW
            ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True
            r+=1
            sub=sub.sort_values(done)           # not-uploaded first
            for _,x in sub.iterrows():
                ws.row_dimensions[r].outlineLevel=2; ws.row_dimensions[r].hidden=True
                if met=='Installs' and USE_REAL_UPLOAD and upl is not None and x['_pid'] in upl.index:
                    _lagval = x['ulag_install_real']   # real timestamp, install rows only
                elif met=='Evals' and EVAL_REAL_ACTIVE and x['_pid'] in eup.index:
                    _lagval = x['ulag_eval_real']       # real timestamp, eval rows -- this branch
                    # never existed before. Eval lag always fell to the proxy here even when the
                    # real eval-upload pull was live, which is exactly why this column and the
                    # Locations tab's "Avg Eval Upload (d)" (which HAS always used real-when-available)
                    # could show materially different numbers for the identical set of projects.
                elif met=='Evals':
                    _lagval = x['ulag']                 # proxy fallback, eval rows without real data
                else:
                    _lagval = float('nan')              # install row, no real pull this run
                # Days to Current -- same dual-anchor convention as the analyst
                # team's own Aging Report: RFQ-received (Created On proxy) for
                # Evals, PO Received Date for Installs. Only meaningful for
                # not-yet-uploaded rows; nothing is aging once it's done.
                _dtc=None; _stage=None
                if not x[done]:
                    _anchor = x.get('cod') if met=='Evals' else x.get('prd')
                    if pd.notna(_anchor):
                        _dtc = (pd.Timestamp.now() - _anchor).total_seconds()/86400
                    _first_call = x.get('First Call Attempt Date and Time') if met=='Evals' else x.get('First Call Attempt Install Date and Time')
                    _calls = x.get('ec') if met=='Evals' else x.get('ic')
                    if pd.isna(_first_call):
                        _stage='Calling (no contact logged yet)'
                    elif pd.isna(_calls) or _calls==0:
                        _stage='Calling'
                    elif not x[done]:
                        _stage='Uploading'
                vals=[met[:-1],x['Project ID'],x.get('Customer'),x[dcol].strftime('%m/%d/%y'),
                      'Yes' if x[done] else 'No',
                      round(_dtc,1) if _dtc is not None else None,
                      _stage,
                      None if pd.isna(x['ec']) else int(x['ec']),
                      None if pd.isna(x['ic']) else int(x['ic']),
                      None if pd.isna(_lagval) else round(float(_lagval),1),
                      None if pd.isna(x['lf']) else int(x['lf']),
                      'Yes' if x.get('_is_reeval') else 'No',
                      clean_note(x.get('Most Recent Timeline Note'))]
                for cc,v in enumerate(vals,1):
                    cel=ws.cell(row=r,column=cc,value=v); cel.font=SM; cel.border=BOX
                    cel.alignment=WRAP if cc==len(DC) else (LFT if cc in (2,3,7) else CTR)
                    if not x[done]: cel.fill=P(LRED3)
                if not x[done]:
                    ws.cell(row=r,column=5).font=Font(name=FN,size=10,bold=True,color=NOUP)
                r+=1

# ==================================================== RE-EVALUATIONS
ws=wb.create_sheet('Re-Evaluations')
ws.sheet_view.showGridLines=False
titles(ws,'Re-Evaluations \u2014 which projects are getting a second visit',
       f'Detected from note text ("RE-EVAL"/"REEVAL") across every project currently OPEN (not sitting '
       f'in Invoicing And Close) for that location -- not the full historical book, and not just this '
       f'week\'s worked population. Per-location tally reads re-evals / total currently-open projects. '
       f'Reason-for-re-eval as its own flagged column is a planned upgrade once note-scanning is available '
       f'at scale -- see the Contact Attempt Days review for why that\'s not routine yet.')
rvr=4
for reg in REGIONS+['Unassigned']:
    locs=set(loc_region.get(reg,[]))
    if not locs: continue
    _reg_reeval_locs={}
    for loc in sorted(locs):
        _loc_active=a[((a['_estrip']==loc)|(a['_istrip']==loc))&a['_is_open']]
        _n=len(_loc_active); _rn=int(_loc_active['_is_reeval'].sum())
        if _n: _reg_reeval_locs[loc]=(_rn,_n,_loc_active)
    if not any(v[0] for v in _reg_reeval_locs.values()): continue   # skip regions with zero re-evals
    band(ws,rvr,f'\u25c6  {reg}',6); rvr+=1
    hdr(ws,rvr,['Location','Re-evals / Active','','','','',''],[38,16,10,10,10,10,10])
    rvr+=1
    for loc,(rn,n,_sub) in sorted(_reg_reeval_locs.items(),key=lambda kv:-kv[1][0]):
        if not rn: continue   # only list locations that actually have a re-eval
        ws.cell(row=rvr,column=1,value=loc).font=BD
        c=ws.cell(row=rvr,column=2,value=f'{rn} \u2013 {n}'); c.alignment=CTR; c.font=BD
        c.number_format='@'   # text, so Excel doesn't date-parse the ratio
        c.fill=P(KPI_A if rn else GRN_Z)
        for cc in range(1,7): ws.cell(row=rvr,column=cc).border=BOX
        rvr+=1
        # detail rows: every actual re-eval project for this location
        hdr(ws,rvr,['      Project','VAMC','Region','Eval Tech','Install Tech','Most Recent Timeline Note'],
            [16,26,10,26,26,70])
        rvr+=1
        for _,prow in _sub[_sub['_is_reeval']].iterrows():
            ws.cell(row=rvr,column=1,value=prow.get('Project ID')).font=SM
            ws.cell(row=rvr,column=2,value=prow.get('Customer')).font=SM
            ws.cell(row=rvr,column=3,value=prow.get('region')).font=SM
            ws.cell(row=rvr,column=4,value=prow.get('Eval Technician')).font=SM
            ws.cell(row=rvr,column=5,value=prow.get('Install Technician')).font=SM
            nc=ws.cell(row=rvr,column=6,value=clean_note(prow.get('Most Recent Timeline Note')))
            nc.font=SM; nc.alignment=WRAP
            for cc in range(1,7): ws.cell(row=rvr,column=cc).border=BOX
            rvr+=1
        rvr+=1
    rvr+=1
caption(ws,rvr,'Reason-for-re-eval column: not yet built. The note text usually states the reason '
        '("veteran refused," "scope change," "RRAA reconsidered") but reliably pulling that out as its '
        'own flagged column needs the same bulk-notes connector fix already requested for Contact Attempt '
        'Days -- see that review for the full explanation of what\'s blocking it.',6)

# ==================================================== ATTENTION DEEP DIVE
ws=wb.create_sheet('Attention Deep Dive')
titles(ws,'Director Attention — flagged this week',
       "Rule-based flags from the export. Latest status is the coordinator's own timeline note.")
hdr(ws,4,['Project ID','Region','Location','Type','Flag','Latest status note','Coordinator','VAMC'],
    [13,10,28,8,24,66,24,34])
ws.freeze_panes='A5'
r=5; flags=[]
for met,W,tcol,done,callcol in [('Eval',EW,'Eval Technician','edone','ec'),
                                ('Install',IW,'Install Technician','idone','ic')]:
    for _,x in W.iterrows():
        f=[]
        if not x[done]: f.append('not uploaded')
        if pd.notna(x[callcol]) and x[callcol]>=8: f.append(f'{int(x[callcol])} calls')
        note=clean_note(x.get('Most Recent Timeline Note'))
        for kw in ['refus','cancel','hold','reschedul','no answer']:
            if kw in note.lower(): f.append(kw); break
        if f: flags.append((len(f),str(x['Project ID']),x['region'],x.get(tcol),met,
                            '; '.join(f),note[:220],x.get('ProjectCoordinator'),x.get('Customer')))
for _,pid,reg,loc,met,fl,note,coord,vamc in sorted(flags,key=lambda t:(-t[0],t[1]))[:60]:
    for cc,v in enumerate([pid,reg,loc,met,fl,note,coord,vamc],1):
        c=ws.cell(row=r,column=cc,value=v); c.font=SM; c.border=BOX
        c.alignment=WRAP if cc==6 else LFT
    ws.cell(row=r,column=5).fill=P(LRED3)
    r+=1

# ==================================================== PACE & DEVIATION
ws=wb.create_sheet('Pace & Deviation')
titles(ws,'Pace & Deviation — this week against each location\'s own history',
       f'Baseline {BASE_FROM.date()} to {WEEK.date()} ({len(BWEEKS)} weeks). Directional bands: '
       'green above Pace, yellow/red below. Overshooting Pace is a good outcome.')
c=ws.cell(row=2,column=8,value='Zone key:'); c.font=BD
for j,(lbl,fl) in enumerate([('Well ahead',GRN2_Z),('Ahead',GRN_Z),('On pace',NEU_Z),
                             ('Caution',YEL_Z),('ALERT',RED_Z)],9):
    c=ws.cell(row=2,column=j,value=lbl); c.fill=P(fl); c.font=SM; c.alignment=CTR; c.border=BOX
    ws.column_dimensions[L(j)].width=12
caption(ws,3,'Full detail behind the Pace columns on the Locations tab. One row per location per metric.',11)
hdr(ws,4,['Location','Metric','Tier','Weeks','Pace/wk','σ','Lower −2σ','Upper +2σ',
          'This week','Zone','Consistency'],[38,9,7,8,11,8,11,11,11,13,13])
ws.freeze_panes='A5'
r=5
PACEROWS={}
for reg in REGIONS+['Unassigned']:
    locs=set(loc_region.get(reg,[]))   # dominant/parent region always -- an
    # outlier project (e.g. one Biloxi eval on an otherwise-East location) stays
    # in that location's own numbers, it just doesn't relist the location itself
    # under a different region band.
    if not locs: continue
    band(ws,r,f'◆  {reg}',11); r+=1
    for loc in sorted(locs):
        for metric,W,tcol,done in [('Install',IW,'Install Technician','idone'),
                                   ('Eval',EW,'Eval Technician','edone')]:
            st=pace_stats(loc,metric)
            if st is None: continue
            this=int(W[W[tcol].astype(str).str.strip()==loc][done].sum())
            zname,zfill=zone_of(this,st)
            vals=[loc,metric,st['tier'],st['n'],round(st['mean'],1),round(st['sd'],1),
                  round(st['lo2'],1),round(st['hi2'],1),this,zname,
                  round(st['cons'],2) if st['cons'] else None]
            if st['tier']=='C': vals[6]=vals[7]=vals[9]='—'
            for cc,v in enumerate(vals,1):
                cel=ws.cell(row=r,column=cc,value=v if v is not None else '—')
                cel.font=BD if cc in (1,9) else RG; cel.border=BOX
                cel.alignment=LFT if cc==1 else CTR
                if cc in (5,6,7,8) and isinstance(v,float): cel.number_format='0.0'
                if cc==11 and isinstance(v,float): cel.number_format='0.00'
            ws.cell(row=r,column=2).fill=P(LRED3 if metric=='Eval' else LBLU3)
            if zfill: ws.cell(row=r,column=10).fill=P(zfill)
            if st['cons'] and st['cons']>1.5: ws.cell(row=r,column=11).fill=P(KPI_A)
            PACEROWS[(loc,metric)]=st
            r+=1
    r+=1

# ==================================================== REPLACEMENT
ws=wb.create_sheet('Replacement')
titles(ws,'Replacement — work in vs work out',
       'Are we taking work in as fast as we complete it? Ratio above 1.00 means the queue grew that week. '
       'RFQ intake uses Created On as a proxy — see Method & Notes.')
a['cro']=pd.to_datetime(a['Created On'],errors='coerce').apply(wk_of)
a['pow']=pd.to_datetime(a['PO Received Date'],errors='coerce').apply(wk_of)
caption(ws,3,'Ratio 1.00 means intake exactly replaced completions that week. Above 1.00 the queue GREW; '
        'below 1.00 it drained. Install intake is PO Received; eval intake uses Created On as a proxy.',3+WOW_N)
hdr(ws,4,['Series','Metric','Region']+[w.strftime('%-m/%-d') for w in WEEKS],[30,9,11]+[9]*WOW_N)
r=5
for metric,incol,inlbl,done,wkcol in [('Install','pow','POs received (IN)','idone','iwk'),
                                      ('Eval','cro','RFQs received (IN, proxy)','edone','ewk')]:
    band(ws,r,f'◆  {metric}',3+WOW_N); r+=1
    rr={}
    for lbl,kind in [(inlbl,'in'),(f'{metric}s completed (OUT)','out'),
                     ('Net (in − out)','net'),('Ratio (in ÷ out)','ratio')]:
        for reg in REGIONS:
            ws.cell(row=r,column=1,value=lbl).font=BD if kind in ('in','out') else RG
            ws.cell(row=r,column=2,value=metric).font=RG
            ws.cell(row=r,column=3,value=reg).font=RG
            for j,w in enumerate(WEEKS,4):
                if kind=='in':
                    v=int(((a[incol]==w)&(a['region']==reg)).sum())
                elif kind=='out':
                    v=int(a[(a[wkcol]==w)&(a['region']==reg)][done].sum())
                elif kind=='net':
                    v=rr[('in',reg)][j]-rr[('out',reg)][j]
                else:
                    o=rr[('out',reg)][j]
                    v=round(rr[('in',reg)][j]/o,2) if o else None
                cel=ws.cell(row=r,column=j,value=v if v is not None else '—')
                cel.font=RG; cel.alignment=CTR; cel.border=BOX
                if kind=='ratio' and isinstance(v,float):
                    cel.number_format='0.00'
                    cel.fill=P(KPI_A) if v>1.15 else (P(GRN_Z) if v<0.85 else P(NEU_Z))
                rr.setdefault((kind,reg),{})[j]=v if isinstance(v,(int,float)) else 0
            for cc in (1,2,3): ws.cell(row=r,column=cc).border=BOX
            r+=1
    r+=1

# ==================================================== TIME ON SITE
ws=wb.create_sheet('Time on Site')
TOS_LO,TOS_HI,TOS_MIN_N,TOS_MIN_COV=15,720,20,0.20
titles(ws,'Time on Site (minutes)',
       f'Values outside {TOS_LO}–{TOS_HI} min excluded as entry errors. An average shows only where at least '
       f'{TOS_MIN_N} clean records exist AND coverage is at least {int(TOS_MIN_COV*100)}%. The field is not '
       'mandatory in AIMS, so most locations read "insufficient".')
a['etos']=pd.to_numeric(a['Eval Time On Site (min)'],errors='coerce')
a['itos']=pd.to_numeric(a['Install Time On Site (min)'],errors='coerce')
hdr(ws,4,['Location','Metric','Region','Worked (baseline)','Logged','Coverage','Avg min','Excluded outliers'],
    [38,9,11,17,10,11,11,16])
ws.freeze_panes='A5'
r=5
for reg in REGIONS+['Unassigned']:
    locs=set(loc_region.get(reg,[]))   # dominant/parent region always -- an
    # outlier project (e.g. one Biloxi eval on an otherwise-East location) stays
    # in that location's own numbers, it just doesn't relist the location itself
    # under a different region band.
    if not locs: continue
    band(ws,r,f'◆  {reg}',8); r+=1
    for loc in sorted(locs):
        for metric,tcol,wkcol,tos in [('Install','Install Technician','iwk','itos'),
                                      ('Eval','Eval Technician','ewk','etos')]:
            sub=a[(a[tcol].astype(str).str.strip()==loc)&(a[wkcol].isin(BWEEKS))]
            if not len(sub): continue
            raw=sub[tos].dropna(); raw=raw[raw>0]
            cl=raw[(raw>=TOS_LO)&(raw<=TOS_HI)]
            cov=len(raw)/len(sub)
            ok=len(cl)>=TOS_MIN_N and cov>=TOS_MIN_COV
            vals=[loc,metric,reg,len(sub),len(raw),cov,
                  round(float(cl.mean())) if ok else 'insufficient',len(raw)-len(cl)]
            for cc,v in enumerate(vals,1):
                cel=ws.cell(row=r,column=cc,value=v); cel.border=BOX
                cel.font=BD if cc==1 else (NOTE if v=='insufficient' else RG)
                cel.alignment=LFT if cc==1 else CTR
                if cc==6: cel.number_format='0%'
            ws.cell(row=r,column=2).fill=P(LRED3 if metric=='Eval' else LBLU3)
            if not ok: ws.cell(row=r,column=7).fill=P(LRED3)
            elif cov<0.5: ws.cell(row=r,column=6).fill=P(KPI_A)
            r+=1
    r+=1

# ==================================================== BOOTS ON GROUND VISUALIZATION
ws=wb.create_sheet('Boots on Ground Visualization')
ws.sheet_view.showGridLines=False
BGLBLCOL=30
ws.column_dimensions[L(BGLBLCOL)].hidden=True
titles(ws,'Boots on Ground Visualization \u2014 receipt to crew on site',
       f'Week of {WEEK.strftime("%b %-d")}. Eval BOG = Eval Date minus RFQ received (Created On '
       'proxy). Install BOG = Install Date minus PO Received Date. Ranked worst first (longest) '
       'by average days -- more days means veterans are waiting longer between request and boots '
       'actually on site.')
bgrow=4
def bog_window(region,metric):
    """VA scoring windows, not an arbitrary threshold. Section 6 of the GAP
    schema: East Eval 5 calendar days, West Eval 3 BUSINESS days (weekends
    don't count against West), Install 10 calendar days everywhere. Central
    and Unassigned follow East, same convention already used for Contact
    Attempt Days thresholds."""
    if metric=='Eval':
        if region=='West': return 3,'business'
        return 5,'calendar'
    return 10,'calendar'   # Install: same window East/Central/West

for metric,W,tcol,bogcol,anchorcol,endcol,colr in [
        ('Eval',EW,'Eval Technician','ebog_adj','cod','edt',AA_RED),
        ('Install',IW,'Install Technician','ibog_adj','prd','idt',AA_BLUE)]:
    band(ws,bgrow,f'\u25c6  {metric} Boots on Ground by location',5); bgrow+=1
    caption(ws,bgrow,(f'Avg BOG = mean days from {"Created On (authoritative eval anchor for this report)" if metric=="Eval" else "PO Received Date (authoritative install anchor)"} '
             f"to {metric} Date, over this week's worked {metric.lower()}s at that location. Day-counts are "
             f"HOLIDAY-ADJUSTED -- the 6 AA holidays are skipped in every mode, per the GAP Late-Codes framework. "
             f"Colored against the actual VA scoring window: East {'5 calendar days' if metric=='Eval' else '10 calendar days'}, "
             f"West {'3 BUSINESS days (weekends and holidays excluded -- the only business-day section)' if metric=='Eval' else '10 calendar days'}, Central follows East. "
             'Chart shows the worst 15 -- click it and use the funnel filter icon to bring any other location into view.'),6)
    bgrow+=1
    bghdrow=bgrow
    ws.cell(row=bghdrow,column=BGLBLCOL,value='(chart axis labels)').font=NOTE
    hdr(ws,bghdrow,[f'{metric} location','Worked','Avg BOG (d)','Min BOG (d)','Max BOG (d)'],[46,11,12,12,12])
    ws.column_dimensions['F'].width=3
    bgrows=[]
    for loc in ALLLOC_MASTER:
        sub=W[W[tcol].astype(str).str.strip()==loc]
        # bogcol is already the holiday-adjusted, mode-correct count (West eval
        # rows carry the business-day value, everything else calendar), so the
        # displayed number and the colour now use the SAME figure -- the
        # framework requires every number in a string to agree.
        bog=pd.to_numeric(sub[bogcol],errors='coerce').dropna() if len(sub) else pd.Series(dtype=float)
        region=LOC_TO_REGION.get(loc,'East')
        window,mode=bog_window(region,metric)
        color_val=float(bog.mean()) if len(bog) else None
        # No row for this side this week at all -- still show the location,
        # with a dash, rather than silently omitting it. -999 sorts these to
        # the bottom of a worst-first (reverse) sort without being confused
        # for a real, excellent (near-zero-day) result.
        bgrows.append((float(bog.mean()) if len(bog) else -999, loc, len(sub),
                      round(float(bog.mean()),1) if len(bog) else None,
                      round(float(bog.min()),1) if len(bog) else None,
                      round(float(bog.max()),1) if len(bog) else None,
                      color_val, window))
    bgrows.sort(reverse=True)
    bgr=bghdrow+1
    for avgbog,loc,tot,avgv,minv,maxv,color_val,window in bgrows:
        ws.cell(row=bgr,column=1,value=loc).font=RG
        ws.cell(row=bgr,column=2,value=tot).alignment=CTR
        c=ws.cell(row=bgr,column=3,value=avgv if avgv is not None else '\u2014'); c.alignment=CTR
        if isinstance(avgv,float) and color_val is not None:
            c.font=BD; c.fill=P(KPI_R) if color_val>window*1.4 else (P(KPI_A) if color_val>window else P(GRN_Z))
        ws.cell(row=bgr,column=4,value=minv if minv is not None else '\u2014').alignment=CTR
        ws.cell(row=bgr,column=5,value=maxv if maxv is not None else '\u2014').alignment=CTR
        ws.cell(row=bgr,column=BGLBLCOL,value=short(loc)).font=SM
        for cc in range(1,6): ws.cell(row=bgr,column=cc).border=BOX
        bgr+=1
    _bgvalid=[t for t in bgrows if t[3] is not None]
    if _bgvalid:
        _bn_chart=min(len(_bgvalid),15)
        bgch=BarChart(); bgch.type='col'; bgch.height=13; bgch.width=21
        bgch.title=f'Avg {metric.lower()} Boots on Ground (days) by location'
        bgch.add_data(Reference(ws,min_col=3,min_row=bghdrow,max_row=bghdrow+_bn_chart),titles_from_data=True)
        _bgcatref=AxDataSource(strRef=StrRef(
            f"'Boots on Ground Visualization'!$A${bghdrow+1}:$A${bghdrow+_bn_chart}"))
        bgch.series[0].cat=_bgcatref
        bgch.style = 4 if metric=='Eval' else 3
        bgch.series[0].graphicalProperties=GraphicalProperties(
            solidFill=ColorChoice(schemeClr=SchemeColor(val='accent2' if metric=='Eval' else 'accent1', shade=38000)))
        bgch.dLbls=DataLabelList(); bgch.dLbls.showVal=True
        bgch.dLbls.showSerName=False; bgch.dLbls.showCatName=False; bgch.dLbls.showLegendKey=False
        bgch.legend=None
        bgch.y_axis.majorGridlines=ChartLines()
        vertical_labels(bgch.x_axis,700)
        ws.add_chart(bgch,f'H{bghdrow}')
    bgrow=bgr+2

# ==================================================== UPLOAD VISUALIZATION
ws=wb.create_sheet('Upload Visualization')
ws.sheet_view.showGridLines=False
LBLCOL=30                      # unused now (categories point at column A), kept
ws.column_dimensions[L(LBLCOL)].hidden=True
titles(ws,'Upload Visualization \u2014 who has finished their paperwork and who has not',
       f'Week of {WEEK.strftime("%b %-d")}. The table lists every location, worst first. The chart shows '
       'ONLY locations with work still outstanding \u2014 that is the follow-up list.')
row=4
for metric,W,tcol,done,colr in [('Eval',EW,'Eval Technician','edone',AA_RED),
                                ('Install',IW,'Install Technician','idone',AA_BLUE)]:
    nxt = 'quote sent' if metric=='Eval' else 'packet sent'
    band(ws,row,f'\u25c6  {metric}s \u2014 worked this week vs still awaiting {nxt}',5); row+=1
    _lagnote=('Avg upload (days) is the mean gap from evaluation to quote sent \u2014 a PROXY, amber over 3 '
              'days, red over 7.' if metric=='Eval' else
              ('Avg upload (days) is the mean gap from install to the REAL document upload timestamp '
               '(get_document_upload_dates) \u2014 amber over 3 days, red over 7.'
               if (USE_REAL_UPLOAD and upl is not None) else
               'Avg upload (days) is not available for installs this run: no upload-dates pull is attached, '
               'and Install Packet Sent Date is blank on nearly every row in the export.'))
    caption(ws,row,f'Pending means the {metric.lower()} was performed but {nxt} has not been recorded in '
            f'AIMS. Column height in the chart is what the location worked; the red cap is what is still '
            f'pending. Sorted worst first. Chart shows the top 15 -- use the funnel filter icon on the '
            f'chart itself (appears when you click it) to bring any other location into view. {_lagnote}',6)
    row+=1
    hdrow=row
    ws.cell(row=hdrow,column=LBLCOL,value='(unused)').font=NOTE
    hdr(ws,hdrow,[f'{metric} location','Worked',f'{IC_DONE} Completed',f'{IC_PEND} Pending',
                  'Pending %','Avg upload (days)'],[46,11,14,13,12,17])
    ws.column_dimensions['F'].width=3
    rows=[]
    for loc in ALLLOC_MASTER:
        sub=W[W[tcol].astype(str).str.strip()==loc]
        ok=int(sub[done].sum()) if len(sub) else 0
        if not len(sub):
            lag=pd.Series(dtype=float)
        elif metric=='Eval':
            lag=sub['ulag'].dropna()                       # proxy, always
        elif USE_REAL_UPLOAD and upl is not None:
            lag=sub[sub['_pid'].isin(upl.index)]['ulag_install_real'].dropna()   # real, this run
        else:
            lag=pd.Series(dtype=float)
        # No row for this side this week at all -- still show the location
        # with a dash, sorted to the bottom (-999), rather than reading as a
        # falsely perfect "0 pending, all done."
        rows.append((len(sub)-ok if len(sub) else -999,len(sub),loc,ok,
                     round(float(lag.mean()),1) if len(lag) else None))
    rows.sort(reverse=True)                       # pending desc, then volume
    r2=hdrow+1
    chart_from=r2
    n_pend=0
    for pend,tot,loc,ok,avg in rows:
        ws.cell(row=r2,column=1,value=loc).font=BD if pend>0 else RG
        # axis labels now come straight from column A (see chart cat ref below)
        ws.cell(row=r2,column=1).alignment=LFT
        for cc,v in ((2,tot),(3,ok),(4,pend if pend>=0 else 0)):
            cel=ws.cell(row=r2,column=cc,value=v); cel.alignment=CTR
            cel.font=BD if (cc==4 and pend>0) else RG
        pr=pend/tot if tot else 0
        c=ws.cell(row=r2,column=5,value=pr); c.number_format='0%'; c.alignment=CTR
        if pend>0:
            c.fill=P(KPI_A if pr<0.30 else KPI_R)
            ws.cell(row=r2,column=4).fill=P(KPI_A if pr<0.30 else KPI_R)
            n_pend+=1
        ac=ws.cell(row=r2,column=6,value=avg if avg is not None else IC_NONE)
        ac.alignment=CTR; ac.number_format='0.0'; ac.fill=P(T_UPLOAD)
        if avg is None: ac.font=NOTE
        elif avg>7: ac.fill=P(KPI_R); ac.font=BD
        elif avg>3: ac.fill=P(KPI_A)
        for cc in range(1,7): ws.cell(row=r2,column=cc).border=BOX
        r2+=1
    # chart: vertical columns, abbreviated names, only locations with work outstanding
    if n_pend:
        n_chart=min(n_pend,15)   # cap so labels have room; the funnel filter icon on
                                 # the chart (native Excel, always there when a chart
                                 # is selected) lets you bring the rest back into view
        ch=BarChart(); ch.type='col'; ch.grouping='stacked'; ch.overlap=100; ch.gapWidth=60
        ch.height=13; ch.width=21   # matches the size set by hand in the reviewed copy
        ch.y_axis.delete=True                          # data labels already show the values
        ch.add_data(Reference(ws,min_col=3,max_col=4,min_row=hdrow,max_row=chart_from+n_chart-1),
                    titles_from_data=True)
        # openpyxl writes a numeric reference here, which Excel cannot render as text
        # labels. Force a string reference so the names actually appear. Points at
        # column 1 -- always visible, never hidden -- rather than a hidden helper
        # column: some Excel builds will not source axis category text from a
        # hidden cell even with plotVisOnly=0, which is why labels vanished before.
        _catref=AxDataSource(strRef=StrRef(
            f"'Upload Visualization'!$A${chart_from}:$A${chart_from+n_chart-1}"))
        for _ser in ch.series: _ser.cat=_catref
        # exact colors/style copied from the manually-styled copy: Excel's built-in
        # Chart Style 4 (Eval, theme accent2) / Style 3 (Install, theme accent1)
        _theme_accent = 'accent2' if metric=='Eval' else 'accent1'
        ch.style = 4 if metric=='Eval' else 3
        ch.series[0].graphicalProperties=GraphicalProperties(
            solidFill=ColorChoice(schemeClr=SchemeColor(val=_theme_accent, shade=76000)))
        ch.series[1].graphicalProperties=GraphicalProperties(
            solidFill=ColorChoice(schemeClr=SchemeColor(val=_theme_accent, tint=77000)))
        ch.dLbls=DataLabelList(); ch.dLbls.showVal=True
        ch.dLbls.showSerName=False; ch.dLbls.showCatName=False; ch.dLbls.showLegendKey=False
        ch.legend.position='b'
        tot_pend=sum(x[0] for x in rows)
        ch.title=f'{metric}s \u2014 completed vs pending'
        ch.y_axis.title=None
        ch.gapWidth=40
        ch.y_axis.majorGridlines=ChartLines()
        vertical_labels(ch.x_axis)          # names read straight down under each column
        ch.legend.position='b'
        ws.add_chart(ch,f'H{hdrow}')
    row=r2+2

# ==================================================== CALLS VISUALIZATION
ws=wb.create_sheet('Calls Visualization')
ws.sheet_view.showGridLines=False
CLBLCOL=30                     # unused now (categories point at column A), kept
ws.column_dimensions[L(CLBLCOL)].hidden=True
titles(ws,'Calls Visualization \u2014 diligence in reaching veterans',
       f'Week of {WEEK.strftime("%b %-d")}. Ranked lowest first by average call attempts \u2014 '
       'more calls means more effort put into reaching that veteran. Low numbers are the ones worth a look.')
crow=4
for metric,W,tcol,callcol,rfqcol,firstcol,colr in [
        ('Eval',EW,'Eval Technician','ec','cod','First Call Attempt Date and Time',AA_RED),
        ('Install',IW,'Install Technician','ic','prd','First Call Attempt Install Date and Time',AA_BLUE)]:
    band(ws,crow,f'\u25c6  {metric} call attempts by location',6); crow+=1
    caption(ws,crow,(f"Avg calls = mean Number Of Call Attempts {metric} over this week's worked " +
             f"{metric.lower()}s at that location. More calls means more diligence in reaching that veteran -- high is good here, low and falling is the pattern worth a look. Avg days to 1st contact is the OPPOSITE polarity -- it's how long after {'RFQ received' if metric=='Eval' else 'PO received'} the first call attempt happened, so low is good there and high is the flag. Chart shows the lowest 15 by avg calls -- click it and use the funnel filter icon to bring any other location into view."),6)
    crow+=1
    chdrow=crow
    ws.cell(row=chdrow,column=CLBLCOL,value='(chart axis labels)').font=NOTE
    hdr(ws,chdrow,[f'{metric} location','Worked','Total calls','Avg calls','Avg days to 1st contact'],[46,11,12,11,17])
    ws.column_dimensions['F'].width=3
    crows=[]
    for loc in ALLLOC_MASTER:
        sub=W[W[tcol].astype(str).str.strip()==loc]
        calls=sub[callcol].dropna() if len(sub) else pd.Series(dtype=float)
        _fc=pd.to_datetime(sub[firstcol],errors='coerce') if len(sub) else pd.Series(dtype='datetime64[ns]')
        _fclag=(_fc-sub[rfqcol]).dt.total_seconds()/86400 if len(sub) else pd.Series(dtype=float)
        _fclag=_fclag.dropna()
        # No row for this side this week at all -- still show the location
        # with a dash rather than omitting it. 999 sorts these to the bottom
        # of an ascending (lowest-first) sort without reading as "0 calls,
        # worst possible."
        crows.append((float(calls.mean()) if len(calls) else 999, loc, len(sub),
                      int(calls.sum()) if len(calls) else 0,
                      round(float(calls.mean()),1) if len(calls) else None,
                      round(float(_fclag.mean()),1) if len(_fclag) else None))
    crows.sort()   # ascending -- lowest average calls first, low is now the flag, not high
    cr=chdrow+1
    for avgc,loc,tot,totcalls,avgv,fcavg in crows:
        ws.cell(row=cr,column=1,value=loc).font=RG
        ws.cell(row=cr,column=2,value=tot).alignment=CTR
        ws.cell(row=cr,column=3,value=totcalls).alignment=CTR
        c=ws.cell(row=cr,column=4,value=avgv if avgv is not None else '\u2014'); c.alignment=CTR
        if isinstance(avgv,float):
            # Inverted from the original: more calls = more diligence = good.
            # Low average calls is the pattern worth flagging now, not high.
            c.font=BD; c.fill=P(KPI_R) if avgv<3 else (P(KPI_A) if avgv<5 else P(GRN_Z))
        fc=ws.cell(row=cr,column=5,value=fcavg if fcavg is not None else '\u2014'); fc.alignment=CTR
        if isinstance(fcavg,float):
            # Opposite polarity from Avg calls -- here LOW is good (fast first
            # outreach), HIGH is the flag (slow to make the first attempt).
            fc.font=BD; fc.fill=P(GRN_Z) if fcavg<=1 else (P(KPI_A) if fcavg<=3 else P(KPI_R))
        ws.cell(row=cr,column=CLBLCOL,value=short(loc)).font=SM
        for cc in range(1,6): ws.cell(row=cr,column=cc).border=BOX
        cr+=1
    _valid=[t for t in crows if t[4] is not None]
    if _valid:
        _n_chart=min(len(_valid),15)   # cap for label room; use the chart's own
                                        # native funnel filter icon to see the rest
        cch=BarChart(); cch.type='col'; cch.height=13; cch.width=21
        cch.title=f'Avg {metric.lower()} call attempts by location'
        cch.add_data(Reference(ws,min_col=4,min_row=chdrow,max_row=chdrow+_n_chart),titles_from_data=True)
        _catref=AxDataSource(strRef=StrRef(
            f"'Calls Visualization'!$A${chdrow+1}:$A${chdrow+_n_chart}"))
        cch.series[0].cat=_catref
        # exact colors/style from the manually-styled copy: Style 4/accent2 for
        # eval, Style 3/accent1 for install, single-series shade 38000
        cch.style = 4 if metric=='Eval' else 3
        cch.series[0].graphicalProperties=GraphicalProperties(
            solidFill=ColorChoice(schemeClr=SchemeColor(val='accent2' if metric=='Eval' else 'accent1', shade=38000)))
        cch.dLbls=DataLabelList(); cch.dLbls.showVal=True
        cch.dLbls.showSerName=False; cch.dLbls.showCatName=False; cch.dLbls.showLegendKey=False
        cch.legend=None
        cch.y_axis.majorGridlines=ChartLines()
        vertical_labels(cch.x_axis,700)
        ws.add_chart(cch,f'I{chdrow}')
    crow=cr+2

# ==================================================== METHOD & NOTES# ==================================================== METHOD & NOTES# ==================================================== METHOD & NOTES
ws=wb.create_sheet('Method & Notes')
ws.column_dimensions['A'].width=112
_basis_txt = ('INSTALLS and EVALS both use REAL document upload timestamps from '
              'get_document_upload_dates -- installs from the cw_installdateandtime pull, '
              'evals from the separate cw_evaldateandtime pull. Both are required; either '
              'missing falls back to its proxy for that side only.'
              if (USE_REAL_UPLOAD and upl is not None and EVAL_REAL_ACTIVE) else
              ('INSTALLS use REAL document upload timestamps; EVALS are on the Quote Sent '
               'Date proxy because no eval-side upload pull (AA_EVAL_UPLOAD_CSV) was attached '
               'this run.'
               if (USE_REAL_UPLOAD and upl is not None) else
               'Quote Sent Date (evals) / Install Packet Sent (installs) -- a proxy, not upload timing'))
lines=[('Method & Notes — how each number is produced',TITLE15),('',RG),
 ('GAP LATE-CODES FRAMEWORK \u2014 clock anchors, day-counts, and one documented divergence',BD),
 ('   HOLIDAYS: the 6 AA holidays are skipped in EVERY mode, per the framework\'s "NON-WORKING',RG),
 ('   DAYS -- the 6 AA holidays, skipped in EVERY mode." Fixed-date holidays (New Year, Jul 4,',RG),
 ('   Christmas) use the observed-day shift (Sat -> preceding Fri, Sun -> following Mon); the',RG),
 ('   three floating holidays never shift. Calendar mode skips holidays ONLY, not weekends.',RG),
 ('   West Eval is the single business-day section (weekends AND holidays skipped).',RG),
 ('   INSTALL ANCHOR: PO Received Date -- confirmed authoritative by Micah, 13 Aug 2026.',RG),
 ('   EVAL ANCHOR: Created On -- confirmed authoritative FOR THIS REPORT by Micah, 13 Aug 2026,',RG),
 ('   as the point the request is confirmed in AA\'s hands and processes.',RG),
 ('   DOCUMENTED DIVERGENCE: the framework states "EVAL ANCHOR = RFQ Received. Currently in',RG),
 ('   force... Do NOT score from Created On until that lands and the engine is updated." This',RG),
 ('   report anchors eval day-counts on Created On anyway, by direction. Until the GAP-Tooling',RG),
 ('   data-governance sweep lands, the same project can show a different eval day-count here than',RG),
 ('   in the late-code engine. This is an authorised, recorded divergence -- not a bug.',RG),
 ('   ENGINE NOTE: the framework says an authoritative engine (sections.py / daycount.py) wins',RG),
 ('   over any restatement of its values. Those files are not available to this builder, so the',RG),
 ('   day-count logic here is implemented from the framework document\'s own Table 0 / Table 1.',RG),
 ('   If the engine becomes available it should be called directly instead of this.',RG),
 ('',RG),
 ('ICONOGRAPHY \u2014 the same symbol means the same thing on every tab',BD),
 ('   \u25cf red = Evals                    \u25cf blue = Installs',RG),
 ('   \u2713 completed                     \u25cb pending / awaiting the next step',RG),
 ('   \u25a0 Service Center                \u25a1 Dealer',RG),
 ('   \u25b2\u25b2 more than 2 sigma ABOVE Pace     \u25b2 1-2 sigma above',RG),
 ('   \u25cf within 1 sigma \u2014 on Pace            \u25bc 1-2 sigma BELOW Pace',RG),
 ('   \u25bc\u25bc more than 2 sigma BELOW Pace \u2014 act on this one',RG),
 ('   \u2191 queue growing   \u2192 level   \u2193 queue draining',RG),
 ('   \u2014 not measured (the field is not populated in AIMS)',RG),
 ('',RG),
 ('MEASURE TINTS \u2014 the same pale colour marks the same kind of number on every tab',BD),
 ('   rose = evaluations       pale blue = installations       lavender = lineal feet',RG),
 ('   peach = call attempts    teal = upload / turnaround       sage = time on site',RG),
 ('   Status fills (green / amber / red) always sit on top of these.',RG),
 ('',RG),
 (f'Scope: AA Weekly Snapshot Report export. Reporting week {WEEK.date()} to {(WEEK+pd.Timedelta(days=6)).date()}.',RG),
 ('Cancelled projects (any non-blank Cancellation Code) are excluded everywhere.',RG),
 (f'COMPLETION BASIS THIS RUN: {_basis_txt}',BD),('',RG),
 ('Eval completed (\u2713)   = Eval Date in the week AND Quote Sent Date present. Else pending.',RG),
 ('                       This is a PROXY. Always in effect for evals -- see the basis line above.',RG),
 ('Install completed (\u2713)= EITHER of two definitions depending on the basis line above:',RG),
 ('   if REAL upload dates: Install Date in the week AND a real install-packet upload timestamp exists.',RG),
 ('   if PROXY (no upload pull attached): Install Date in the week AND Install Packet Sent = Yes.',RG),
 ('                       Check the basis line above for which one produced the numbers on this run.',RG),
 ("Total LF             = SUM 'Linear Feet (CadQuoting)' over install completions where entered",RG),
 ('                       (blanks skipped, never counted as 0).',RG),
 ('Avg Call Attempts    = mean of populated Number Of Call Attempts Eval / Install.',RG),
 ('Avg Upload (d)       = mean of (Quote Sent Date \u2212 Eval Date). Install Packet Sent Date is',RG),
 ('                       blank on all but 4 rows in the export, so this is eval-side only.',RG),
 ('Eval quote-sent (d)  = SAME metric as above, renamed for honesty: it measures how fast a',RG),
 ('                       quote follows an eval, NOT a document upload. Always a PROXY.',RG),
 ('Install upload (d)   = mean of (REAL install-packet upload timestamp \u2212 Install Date),',RG),
 ('                       timezone-localized per project state. Only populated when an',RG),
 ('                       upload-dates pull is attached for this week; reads \"n/a\" otherwise.',RG),
 ('                       Never falls back to a proxy, so it cannot be mistaken for the row above.',RG),
 ("Region               = the VAMC's VISN (trailing #N or VISN N), per project.",RG),
 ('                       East 1-8 · Central 9,10,11,12,15,16,23 · West 17,19,20,21,22',RG),
 ("Channel              = 'Service Center' if the technician name contains it, else Dealer.",RG),
 ('Director Attention   = not uploaded, or 8+ call attempts, or a timeline note mentioning',RG),
 ('                       refusal / cancellation / hold / reschedule / no answer.',RG),('',RG),
 ('',RG),
 ('Pace/wk             = MEAN of that location\'s weekly completions across the baseline',RG),
 ('                      (first week it appears onward; later empty weeks count as 0).',RG),
 ('sigma               = STDEV.S of the same weekly counts.',RG),
 ('Band                = Pace ± 2 sigma, lower bound floored at zero.',RG),
 ('Zone                = DIRECTIONAL. Above Pace is good: green beyond 1 sigma, deeper green',RG),
 ('                      beyond 2 sigma. Below Pace: yellow beyond 1 sigma, red beyond 2 sigma.',RG),
 ('                      Grey = within 1 sigma either way. About a third of weeks land beyond',RG),
 ('                      1 sigma by chance, so yellow means LOOK, red means ACT.',RG),
 ('Tier                = A: 12+ weeks and Pace >= 0.92/wk, own sigma.  B: 6+ weeks, pooled.',RG),
 ('                      C: under 6 weeks, no band shown.',RG),
 ('Consistency         = sigma / sqrt(Pace). Near 1.0 is ordinary count noise; above 1.5 is',RG),
 ('                      genuinely erratic. Analyst column, not for supervisor packets.',RG),
 ('Keep-up rate       = work COMPLETED / work RECEIVED per week. Install intake = PO Received Date.',RG),
 ('                      Eval intake = Created On (PROXY: the export carries no RFQ Received',RG),
 ('                      column; cra55_rfqreceived exists in Dataverse but is not exported).',RG),
 ('Time on Site        = mean of 15-720 min values, shown only at 20+ records and 20% coverage.',RG),
 ('',RG),
 ('Blank Upload / Call / LF values mean the field was never entered in AIMS, not zero.',NOTE),
 ('VISN 11 assigned to Central per Micah, 6 Aug 2026.',NOTE),
 ('FOIA comparison tab removed per Micah, 6 Aug 2026 — quantities were unreliable.',NOTE),
 ('Do NOT run a LibreOffice recalc on this file; it strips the Scoreboard data bars.',NOTE)]
for i,(t,f) in enumerate(lines,1): ws.cell(row=i,column=1,value=t).font=f

# ==================================================== DATA (WEEKS) — hidden
ws=wb.create_sheet('Data (weeks)')
hdr(ws,1,['Week','Region','Channel','Location','Evals ✓','Evals Pending','Installs ✓',
          'Installs Pending','Total LF','Eval quote-sent (d)','Avg Call Attempts'],
    [12,11,15,38,10,13,11,15,10,14,16])
r=2
for w in WEEKS:
    for loc in sorted(allloc):
        e=a[(a['ewk']==w)&(a['Eval Technician'].astype(str).str.strip()==loc)]
        i=a[(a['iwk']==w)&(a['Install Technician'].astype(str).str.strip()==loc)]
        if not len(e) and not len(i): continue
        sub=pd.concat([e,i])
        reg=sub['region'].mode().iat[0] if len(sub['region'].mode()) else 'Unassigned'
        ul=e['ulag'].dropna(); calls=pd.concat([e['ec'],i['ic']]).dropna()
        vals=[w.strftime('%Y-%m-%d'),reg,'Service Center' if is_sc(loc) else 'Dealer',loc,
              int(e['edone'].sum()),len(e)-int(e['edone'].sum()),
              int(i['idone'].sum()),len(i)-int(i['idone'].sum()),
              int(i['lf'].sum()),
              round(float(ul.mean()),1) if len(ul) else None,
              round(float(calls.mean()),1) if len(calls) else None]
        for cc,v in enumerate(vals,1):
            c=ws.cell(row=r,column=cc,value=v); c.font=SM; c.border=BOX
        r+=1
ws.sheet_state='hidden'

# ---- final tab order: exec first, machinery hidden ----
# ---- 6 visible tabs; supporting detail kept but tucked away ----
VISIBLE=['Scoreboard','Locations','Boots on Ground Visualization','Calls Visualization','Upload Visualization','Re-Evaluations','5 Week Trends','6 Month Trends',
         'Completions Detail','Method & Notes']
BACKUP =['Attention Deep Dive','Pace & Deviation','Replacement','Time on Site','Data (weeks)']
for i,n in enumerate(VISIBLE+BACKUP):
    if n in wb.sheetnames: wb.move_sheet(n, offset=i-wb.sheetnames.index(n))
for n in BACKUP:
    if n in wb.sheetnames: wb[n].sheet_state='hidden'
if 'By Location (Weekly)' in wb.sheetnames: del wb['By Location (Weekly)']
wb['Data (weeks)'].sheet_state='veryHidden'

# openpyxl writes category references as numRef even when the cells hold text,
# which makes axis labels render as 1,2,3 or vanish. Every category range in this
# workbook is text (week dates, location names), so force them all to strRef.
_fixed=0
for _sn in wb.sheetnames:
    for _ch in wb[_sn]._charts:
        for _sub in ([_ch]+list(getattr(_ch,'_charts',[]) or [])):
            for _ser in getattr(_sub,'ser',[]) or []:
                _cat=getattr(_ser,'cat',None)
                if _cat is not None and _cat.numRef is not None and _cat.strRef is None:
                    _ser.cat=AxDataSource(strRef=StrRef(_cat.numRef.f)); _fixed+=1
print('category refs forced to text:',_fixed)

# ---- open-ready view: every column wide enough for its content, cursor at A1 ----
def _fit(ws, floor=8, ceil=60):
    """Widen any column whose content overflows. Never narrows a deliberate width."""
    need={}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None: continue
            col=c.column_letter
            if ws.column_dimensions[col].hidden: continue
            txt=str(c.value)
            if txt.startswith('='): continue                 # formulas, not display text
            w=max((len(t) for t in txt.split('\n')), default=0)
            if getattr(c.font,'bold',False): w=int(w*1.08)+1
            if w>need.get(col,0): need[col]=w
    for col,w in need.items():
        cur=ws.column_dimensions[col].width or 0
        tgt=min(ceil,max(floor,w+2))
        if tgt>cur: ws.column_dimensions[col].width=tgt

for _sn in wb.sheetnames:
    _ws=wb[_sn]
    if _ws.sheet_state!='visible': continue
    _fit(_ws)
    _lastc=max(_ws.max_column or 1, 8)
    _already=set()
    for _rng in list(_ws.merged_cells.ranges):
        if _rng.min_row<=4 and _rng.min_col==1:
            for _rr2 in range(_rng.min_row,_rng.max_row+1): _already.add(_rr2)
    for _rr in (1,2,3,4):
        if _rr in _already: continue          # caption()/titles() already merged this row
        _v=_ws.cell(row=_rr,column=1).value
        _v2=_ws.cell(row=_rr,column=2).value
        # only merge a genuine single-cell title. A header/data row (hdr(), band())
        # always has a second column populated -- merging one of those wipes every
        # cell but the anchor, which is exactly how the Locations/Trends/Upload
        # Visualization/Completions Detail header rows lost their text.
        if _v and _v2 is None:
            try:
                _ws.merge_cells(start_row=_rr,start_column=1,end_row=_rr,end_column=min(_lastc,20))
            except Exception:
                pass
    _ws.sheet_view.zoomScale=90
    _ws.sheet_view.selection[0].activeCell='A1'
    _ws.sheet_view.selection[0].sqref='A1'
wb.active=0

wb.save(OUT)

# Excel will not plot data from hidden rows or columns unless plotVisOnly is 0.
# The per-location series live in hidden columns, so without this every one of
# those charts opens EMPTY. openpyxl does not expose the flag, so patch the XML.
#
# Also: without <autoTitleDeleted val="1"/>, Excel does not know the manual
# <title> should REPLACE its own auto-generated one -- some Excel builds then
# draw both stacked on top of each other, producing garbled overlapping text
# exactly where the title sits. openpyxl never writes this flag, so patch it in
# alongside plotVisOnly rather than as a separate pass.
def _fix_picture_xfrm(path):
    """REBUILD any drawing part containing a picture, from scratch, in Excel's
    exact dialect.

    Why rebuild rather than patch: Excel's repair log named this exact part --
    "Removed Part: /xl/drawings/drawing1.xml part. (Drawing shape)" -- and
    ground truth for what Excel WANTS came from a copy of this very workbook
    that Excel itself opened and re-saved. That reference file proves:

      * Mixing charts and a picture in ONE drawing part is fine; Excel does it.
      * Excel uses the xdr: PREFIX with xmlns:xdr + xmlns:a on the root.
        openpyxl uses a DEFAULT namespace and re-declares xmlns:a inline
        on every a: element.
      * Excel writes <xdr:graphicFrame macro=""> -- openpyxl omits macro.
      * Excel puts NO <a:xfrm> in a picture's spPr (position comes from the
        anchor) but DOES put <xdr:xfrm> with explicit off/ext on graphicFrames.
      * Excel includes <a:avLst/> inside <a:prstGeom>.
      * Excel's shape ids start at 2.

    Earlier attempts transformed openpyxl's text with regex and left subtle
    gaps (an unprefixed <xfrm/> in no namespace; a wrongly-ADDED picture xfrm).
    Parsing the structure and re-emitting it removes that whole class of bug:
    every byte of the output is written here, in the reference dialect.
    Chart-only drawings are left untouched -- they already open cleanly.
    """
    import zipfile, shutil, re
    from lxml import etree
    XDR='http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    A='http://schemas.openxmlformats.org/drawingml/2006/main'
    R='http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    C='http://schemas.openxmlformats.org/drawingml/2006/chart'
    NS={'x':XDR,'a':A,'r':R,'c':C}

    def _mk(anchors):
        out=['<xdr:wsDr xmlns:xdr="%s" xmlns:a="%s">'%(XDR,A)]
        sid=2                      # Excel starts shape ids at 2
        for kind,frm,to,ext,rid in anchors:
            f=('<xdr:from><xdr:col>%d</xdr:col><xdr:colOff>%d</xdr:colOff>'
               '<xdr:row>%d</xdr:row><xdr:rowOff>%d</xdr:rowOff></xdr:from>')%frm
            if kind=='chart':
                t=('<xdr:to><xdr:col>%d</xdr:col><xdr:colOff>%d</xdr:colOff>'
                   '<xdr:row>%d</xdr:row><xdr:rowOff>%d</xdr:rowOff></xdr:to>')%to
                out.append(
                    '<xdr:twoCellAnchor>'+f+t+
                    '<xdr:graphicFrame macro="">'
                    '<xdr:nvGraphicFramePr>'
                    '<xdr:cNvPr id="%d" name="Chart %d"/>'%(sid,sid-1)+
                    '<xdr:cNvGraphicFramePr/>'
                    '</xdr:nvGraphicFramePr>'
                    '<xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm>'
                    '<a:graphic><a:graphicData uri="%s">'%C+
                    '<c:chart xmlns:c="%s" xmlns:r="%s" r:id="%s"/>'%(C,R,rid)+
                    '</a:graphicData></a:graphic>'
                    '</xdr:graphicFrame>'
                    '<xdr:clientData/>'
                    '</xdr:twoCellAnchor>')
            else:
                out.append(
                    '<xdr:oneCellAnchor>'+f+
                    '<xdr:ext cx="%d" cy="%d"/>'%ext+
                    '<xdr:pic>'
                    '<xdr:nvPicPr>'
                    '<xdr:cNvPr id="%d" name="Image %d" descr="Picture"/>'%(sid,sid-1)+
                    '<xdr:cNvPicPr/>'
                    '</xdr:nvPicPr>'
                    '<xdr:blipFill>'
                    '<a:blip xmlns:r="%s" r:embed="%s" cstate="print"/>'%(R,rid)+
                    '<a:stretch><a:fillRect/></a:stretch>'
                    '</xdr:blipFill>'
                    '<xdr:spPr>'
                    '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                    '</xdr:spPr>'
                    '</xdr:pic>'
                    '<xdr:clientData/>'
                    '</xdr:oneCellAnchor>')
            sid+=1
        out.append('</xdr:wsDr>')
        return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                +''.join(out)).encode('utf-8')

    def _cell(el):
        g=lambda n: int(el.find('x:'+n,NS).text or 0)
        return (g('col'),g('colOff'),g('row'),g('rowOff'))

    tmp=path+'.tmp'; rebuilt=0
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp,'w',zipfile.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            data=zin.read(it.filename)
            if re.match(r'xl/drawings/drawing\d+\.xml$', it.filename):
                root=etree.fromstring(data)
                has_pic=root.find('.//x:pic',NS) is not None
                if has_pic:
                    anchors=[]
                    for anc in root:
                        tag=etree.QName(anc).localname
                        frm=_cell(anc.find('x:from',NS))
                        to=(0,0,0,0); ext=(914400,914400)
                        toel=anc.find('x:to',NS)
                        if toel is not None: to=_cell(toel)
                        extel=anc.find('x:ext',NS)
                        if extel is not None:
                            ext=(int(extel.get('cx')),int(extel.get('cy')))
                        gf=anc.find('x:graphicFrame',NS)
                        if gf is not None:
                            ch=gf.find('.//c:chart',NS)
                            rid=ch.get('{%s}id'%R) if ch is not None else None
                            anchors.append(('chart',frm,to,ext,rid))
                        else:
                            bl=anc.find('.//a:blip',NS)
                            rid=bl.get('{%s}embed'%R) if bl is not None else None
                            anchors.append(('pic',frm,to,ext,rid))
                    data=_mk(anchors); rebuilt+=1
            zout.writestr(it, data)
    shutil.move(tmp,path)
    return rebuilt

def _plot_hidden(path):
    import zipfile, shutil, os, re
    tmp=path+'.tmp'
    n=0; nt=0
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp,'w',zipfile.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            data=zin.read(it.filename)
            if re.match(r'xl/charts/chart\d+\.xml$', it.filename):
                txt=data.decode('utf-8')
                if '<plotVisOnly val="1"/>' in txt:
                    txt=txt.replace('<plotVisOnly val="1"/>','<plotVisOnly val="0"/>'); n+=1
                elif 'plotVisOnly' not in txt:
                    txt=txt.replace('</chart>','<plotVisOnly val="0"/></chart>'); n+=1
                if '<title>' in txt and 'autoTitleDeleted' not in txt:
                    txt=txt.replace('<title>','<title>',1)  # no-op, title stays first
                    txt=txt.replace('</title>','</title><autoTitleDeleted val="1"/>',1); nt+=1
                data=txt.encode('utf-8')
            zout.writestr(it, data)
    shutil.move(tmp,path)
    return n,nt
_pic_fixed=_fix_picture_xfrm(OUT)
_n,_nt=_plot_hidden(OUT)
print('picture xfrm/avLst fixed on', _pic_fixed, 'picture(s)')
print('plotVisOnly patched on', _n, 'charts')
print('autoTitleDeleted patched on', _nt, 'charts')
print('saved',OUT)
print('tabs:',[f'{n}{" (hidden)" if wb[n].sheet_state=="hidden" else ""}' for n in wb.sheetnames])
print(f'week {WEEK.date()}: evals {ec}/{len(EW)} ({ec/max(len(EW),1):.0%})  '
      f'installs {ic}/{len(IW)} ({ic/max(len(IW),1):.0%})  LF {int(lf.sum()):,}  '
      f'avg upload {ulag.mean():.1f}d  flags {len(flags)}')
